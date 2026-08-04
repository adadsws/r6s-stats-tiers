from __future__ import annotations

import argparse
import io
import json
import re
import shutil
import subprocess
import sys
import tempfile
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Mapping, Optional, Tuple
from urllib.parse import urlencode

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PillowImage
from PIL import ImageDraw, ImageFont

from . import report_theme as theme
from .patch_notes import (
    PatchNotesError,
    add_patch_notes_sheet,
)
from .sources import ReportSources, SourceDataError, load_report_sources
from .tiers import SCORE_TO_DISPLAY_TIER, TIER_COLORS, VISIBLE_TIER_ORDER
from .workbook_sources import append_source_footer


SOURCE_SHEETS = ("进攻方", "防守方")
REQUIRED_HEADERS = (
    "图标",
    "干员",
    "速度",
    "主手自动枪械（射速，发/分钟）",
    "副手自动枪械（射速，发/分钟）",
    "主狙",
    "副手霰弹",
    "次要装备",
    "Athieno评分",
)
SCORE_TIERS = SCORE_TO_DISPLAY_TIER
RPM_PATTERN = re.compile(r"（([0-9]+(?:\.[0-9]+)?)）")
GADGET_PATTERN = re.compile(r"^(.+?)(?:[×xX]\s*(\d+))?$")
GADGET_ALIASES = {
    "手雷": "破片手榴弹",
    "破片手榴弹": "破片手榴弹",
    "烟雾手榴弹": "烟雾弹",
    "烟雾弹": "烟雾弹",
    "眩晕手榴弹": "闪光弹",
    "闪光弹": "闪光弹",
}
GADGET_FILES = {
    "倒刺铁丝网": "文件:R6S gp Barbed wire.png",
    "冲击手榴弹": "文件:R6S gp Impact Grenade.png",
    "感应警报器": "文件:R6S gp Proximity Alarm.png",
    "破片手榴弹": "文件:R6S gp Frag Grenade.png",
    "机动护盾": "文件:R6S gp Deployable Shield.png",
    "烟雾弹": "文件:R6S gp Smoke Grenade.png",
    "爆破炸药": "文件:R6S gp Breach Charge.png",
    "电磁脉冲式冲击弹": "文件:R6S gp Impact emp Grenade.png",
    "闪光弹": "文件:R6S gp Stun Grenade.png",
    "硬突破炸药": "文件:R6S gp SecondaryBreacher.png",
    "观测工具阻拦器": "文件:R6S gp Observation Blocker.png",
    "遥控炸药": "文件:R6S gp Nitro Cell.png",
    "阔剑地雷": "文件:R6S gp Claymore.png",
    "防弹摄像头": "文件:R6S gp Bulletproof camera.png",
}
GADGET_DIRECT_URLS = {
    "电磁脉冲式冲击弹": (
        "https://staticctf.ubisoft.com/"
        "J3yJr34U2pZ2Ieem48Dwy9uqj5PNUQTn/"
        "7izurbA5jDmnsmdeBdgKZO/"
        "29bca81243dda4084a92521ac0c03592/"
        "R6S-EMP-Impact-grenade.png"
    ),
}
API_URL = "https://r6s.huijiwiki.com/api.php"
HUJI_IMAGE_PREFIX = "https://huiji-public.huijistatic.com/r6s/"
TIER_ORDER = VISIBLE_TIER_ORDER
CARDS_PER_ROW = 5


class TierChartError(Exception):
    """Raised when source data or required assets are invalid."""


@dataclass(frozen=True)
class WeaponItem:
    name: str
    icon_key: str
    firerate: Optional[int] = None


@dataclass(frozen=True)
class GadgetItem:
    name: str
    quantity: Optional[int]


@dataclass(frozen=True)
class OperatorCard:
    side: str
    name: str
    speed: int
    score: int
    tier: str
    primary_rpms: Tuple[int, ...]
    secondary_rpms: Tuple[int, ...]
    has_semiautomatic: bool
    has_secondary_shotgun: bool
    gadgets: Tuple[GadgetItem, ...]
    source_order: int
    primary_weapons: Tuple[WeaponItem, ...] = ()
    secondary_weapons: Tuple[WeaponItem, ...] = ()
    semiautomatic_weapons: Tuple[WeaponItem, ...] = ()
    secondary_shotguns: Tuple[WeaponItem, ...] = ()


def extract_rpms(value: object) -> Tuple[int, ...]:
    return tuple(
        item.firerate
        for item in parse_automatic_weapons(value)
        if item.firerate is not None
    )


def weapon_icon_key(name: str) -> str:
    normalized = unicodedata.normalize("NFKC", name).casefold()
    normalized = normalized.replace("p10 roni转换套件衍生型", "p10 roni")
    return re.sub(r"[^\w]+", "-", normalized).strip("-")


def parse_automatic_weapons(value: object) -> Tuple[WeaponItem, ...]:
    text = _required_text(value, "自动枪械")
    if text == "无自动枪械":
        return ()
    items = []
    for raw_item in text.splitlines():
        match = re.fullmatch(r"(.+?)（([0-9]+(?:\.[0-9]+)?)）", raw_item.strip())
        if not match:
            raise TierChartError(f"无法提取射速：{raw_item}")
        name, raw_rate = match.groups()
        rate = float(raw_rate)
        if not rate.is_integer():
            raise TierChartError(f"射速必须是整数：{raw_rate}")
        items.append(WeaponItem(name, weapon_icon_key(name), int(rate)))
    return tuple(items)


def parse_named_weapons(value: object) -> Tuple[WeaponItem, ...]:
    text = _required_text(value, "武器状态")
    if text == "无":
        return ()
    return tuple(
        WeaponItem(name, weapon_icon_key(name))
        for name in (line.strip() for line in text.splitlines())
        if name
    )


def parse_gadgets(value: object) -> Tuple[GadgetItem, ...]:
    text = _required_text(value, "次要装备")
    items = []
    for raw_item in re.split(r"[\r\n;；]+", text):
        raw_item = raw_item.strip()
        if not raw_item:
            continue
        match = GADGET_PATTERN.fullmatch(raw_item)
        if not match:
            raise TierChartError(f"无法解析次要装备：{raw_item}")
        name = match.group(1).strip()
        name = GADGET_ALIASES.get(name, name)
        quantity = int(match.group(2)) if match.group(2) else None
        if quantity is not None and quantity <= 0:
            raise TierChartError(f"次要装备数量必须大于零：{raw_item}")
        items.append(GadgetItem(name, quantity))
    if not items:
        raise TierChartError("次要装备为空")
    return tuple(items)


def load_operator_cards(path: Path) -> Dict[str, List[OperatorCard]]:
    source_path = Path(path)
    if not source_path.is_file():
        raise TierChartError(f"找不到输入文件：{source_path}")

    try:
        workbook = load_workbook(source_path, data_only=True)
    except Exception as exc:
        raise TierChartError(f"无法打开输入工作簿 {source_path}：{exc}") from exc

    try:
        missing_sheets = [name for name in SOURCE_SHEETS if name not in workbook.sheetnames]
        if missing_sheets:
            raise TierChartError(f"缺少必需工作表：{', '.join(missing_sheets)}")

        cards: Dict[str, List[OperatorCard]] = {side: [] for side in SOURCE_SHEETS}
        seen_names = set()
        source_order = 0
        for side in SOURCE_SHEETS:
            sheet = workbook[side]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            missing_headers = [header for header in REQUIRED_HEADERS if header not in headers]
            if missing_headers:
                raise TierChartError(
                    f"工作表 {side} 缺少必需表头：{', '.join(missing_headers)}"
                )
            columns = {header: headers.index(header) for header in REQUIRED_HEADERS}
            max_data_row = _operator_data_max_row(sheet, columns)

            for excel_row, values in enumerate(
                sheet.iter_rows(
                    min_row=2,
                    max_row=max_data_row,
                    values_only=True,
                ),
                start=2,
            ):
                name_value = values[columns["干员"]]
                if name_value is None and all(value is None for value in values):
                    continue
                name = _required_text(name_value, f"{side} 第 {excel_row} 行干员")
                if name in seen_names:
                    raise TierChartError(f"重复干员：{name}")
                seen_names.add(name)

                speed = _integer(values[columns["速度"]], f"速度不是数字：{name}")
                score = _integer(values[columns["Athieno评分"]], f"评分不是数字：{name}")
                if score not in SCORE_TIERS:
                    raise TierChartError(f"未知评分：{name} = {score}")

                primary_weapons = parse_automatic_weapons(
                    values[columns["主手自动枪械（射速，发/分钟）"]]
                )
                secondary_weapons = parse_automatic_weapons(
                    values[columns["副手自动枪械（射速，发/分钟）"]]
                )
                semiautomatic_weapons = parse_named_weapons(
                    values[columns["主狙"]]
                )
                secondary_shotguns = parse_named_weapons(
                    values[columns["副手霰弹"]]
                )

                cards[side].append(
                    OperatorCard(
                        side=side,
                        name=name,
                        speed=speed,
                        score=score,
                        tier=SCORE_TIERS[score],
                        primary_rpms=tuple(
                            item.firerate for item in primary_weapons
                        ),
                        secondary_rpms=tuple(
                            item.firerate for item in secondary_weapons
                        ),
                        has_semiautomatic=bool(semiautomatic_weapons),
                        has_secondary_shotgun=bool(secondary_shotguns),
                        gadgets=parse_gadgets(values[columns["次要装备"]]),
                        source_order=source_order,
                        primary_weapons=primary_weapons,
                        secondary_weapons=secondary_weapons,
                        semiautomatic_weapons=semiautomatic_weapons,
                        secondary_shotguns=secondary_shotguns,
                    )
                )
                source_order += 1
        return cards
    finally:
        workbook.close()


def _operator_data_max_row(sheet, columns: Mapping[str, int]) -> int:
    if not sheet.tables:
        return sheet.max_row

    first_required_column = min(columns.values()) + 1
    last_required_column = max(columns.values()) + 1
    matching_rows = []
    for table in sheet.tables.values():
        min_column, min_row, max_column, max_row = range_boundaries(table.ref)
        if (
            min_row == 1
            and min_column <= first_required_column
            and max_column >= last_required_column
        ):
            matching_rows.append(max_row)
    if len(matching_rows) != 1:
        raise TierChartError(
            "工作表 %s 必须有且仅有一个覆盖九列表头的结构化数据表"
            % sheet.title
        )
    return matching_rows[0]


def resolve_wiki_file_url(
    file_title: str,
    query_json: Optional[Callable[[Mapping[str, str]], Mapping[str, object]]] = None,
) -> str:
    requester = query_json or _query_huiji_json
    response = requester(
        {
            "action": "query",
            "titles": file_title,
            "prop": "imageinfo",
            "iiprop": "url",
            "format": "json",
            "formatversion": "2",
        }
    )
    try:
        url = response["query"]["pages"][0]["imageinfo"][0]["url"]
    except (KeyError, IndexError, TypeError):
        raise TierChartError(f"灰机 Wiki 图标地址缺失：{file_title}")
    if not isinstance(url, str) or not url.startswith(HUJI_IMAGE_PREFIX):
        raise TierChartError(f"灰机 Wiki 图标地址不受信任：{file_title}")
    return url


def prepare_gadget_icons(
    items: Iterable[GadgetItem],
    directory: Path,
    *,
    query_json: Optional[Callable[[Mapping[str, str]], Mapping[str, object]]] = None,
    run_command: Callable[..., object] = subprocess.run,
    which: Callable[[str], Optional[str]] = shutil.which,
    sleep: Callable[[float], None] = time.sleep,
) -> Dict[str, Path]:
    unique_names = tuple(dict.fromkeys(item.name for item in items))
    unknown = [name for name in unique_names if name not in GADGET_FILES]
    if unknown:
        raise TierChartError(f"未知次要装备：{', '.join(unknown)}")

    icon_dir = Path(directory)
    icon_dir.mkdir(parents=True, exist_ok=True)
    destinations = {name: icon_dir / _gadget_filename(name) for name in unique_names}
    missing = []
    for name, destination in destinations.items():
        if destination.is_file() and _is_valid_image(destination):
            continue
        destination.unlink(missing_ok=True)
        missing.append((name, destination))
    if not missing:
        return destinations

    curl_path = which("curl.exe") or which("curl")
    if not curl_path:
        raise TierChartError("未找到 curl.exe 或 curl，请先安装并加入 PATH")

    for name, destination in missing:
        url = GADGET_DIRECT_URLS.get(name)
        if url is None:
            url = resolve_wiki_file_url(GADGET_FILES[name], query_json)
        last_error = "未知下载错误"
        for attempt in range(1, 5):
            temporary = destination.with_suffix(".download")
            temporary.unlink(missing_ok=True)
            try:
                run_command(
                    [
                        curl_path,
                        "--location",
                        "--silent",
                        "--show-error",
                        "--fail",
                        "--max-time",
                        "30",
                        "--user-agent",
                        "r6-tier-chart/1.0",
                        "--output",
                        str(temporary),
                        url,
                    ],
                    check=True,
                )
                if not _is_valid_image(temporary):
                    raise ValueError("下载内容不是有效图片")
                if name in GADGET_DIRECT_URLS:
                    with PillowImage.open(temporary) as image:
                        alpha = image.convert("RGBA").getchannel("A")
                        line_art = PillowImage.new(
                            "RGBA",
                            image.size,
                            (0, 0, 0, 0),
                        )
                        line_art.putalpha(alpha)
                        line_art.save(temporary, format="PNG")
                temporary.replace(destination)
                break
            except (OSError, subprocess.CalledProcessError, ValueError) as exc:
                last_error = str(exc)
                temporary.unlink(missing_ok=True)
                if attempt < 4:
                    sleep(float(attempt))
        else:
            raise TierChartError(f"下载 {name} 图标经过 4 次尝试仍失败：{last_error}")
    return destinations


def prepare_weapon_icons(
    items: Iterable[WeaponItem],
    directory: Path,
    *,
    query_json: Optional[
        Callable[[Mapping[str, str]], Mapping[str, object]]
    ] = None,
    run_command: Callable[..., object] = subprocess.run,
    which: Callable[[str], Optional[str]] = shutil.which,
    sleep: Callable[[float], None] = time.sleep,
) -> Dict[str, Path]:
    unique = {}
    for item in items:
        unique.setdefault(item.icon_key, item)

    icon_dir = Path(directory)
    icon_dir.mkdir(parents=True, exist_ok=True)
    destinations = {
        key: icon_dir / (key + ".png")
        for key in unique
    }
    missing = []
    for key, destination in destinations.items():
        if destination.is_file() and _is_valid_image(destination):
            continue
        destination.unlink(missing_ok=True)
        missing.append((unique[key], destination))
    if not missing:
        return destinations

    curl_path = which("curl.exe") or which("curl")
    if not curl_path:
        raise TierChartError("未找到 curl.exe 或 curl，请先安装并加入 PATH")

    for item, destination in missing:
        file_title = "文件:R6S wpn %s.png" % item.name
        url = resolve_wiki_file_url(file_title, query_json)
        last_error = "未知下载错误"
        for attempt in range(1, 5):
            temporary = destination.with_suffix(".download")
            temporary.unlink(missing_ok=True)
            try:
                run_command(
                    [
                        curl_path,
                        "--location",
                        "--silent",
                        "--show-error",
                        "--fail",
                        "--max-time",
                        "30",
                        "--user-agent",
                        "r6-tier-chart/1.0",
                        "--output",
                        str(temporary),
                        url,
                    ],
                    check=True,
                )
                if not _is_valid_image(temporary):
                    raise ValueError("下载内容不是有效图片")
                with PillowImage.open(temporary) as image:
                    source = image.convert("RGBA")
                    bounds = source.getchannel("A").getbbox()
                    if bounds is None:
                        raise ValueError("枪械图标没有可见像素")
                    source = source.crop(bounds)
                    color = bytes.fromhex(theme.COLOURS["text"])
                    silhouette = PillowImage.new(
                        "RGBA",
                        source.size,
                        (color[0], color[1], color[2], 0),
                    )
                    silhouette.putalpha(source.getchannel("A"))
                    silhouette.save(temporary, format="PNG")
                temporary.replace(destination)
                break
            except (
                OSError,
                subprocess.CalledProcessError,
                ValueError,
            ) as exc:
                last_error = str(exc)
                temporary.unlink(missing_ok=True)
                if attempt < 4:
                    sleep(float(attempt))
        else:
            raise TierChartError(
                "下载 %s 图标经过 4 次尝试仍失败：%s"
                % (item.name, last_error)
            )
    return destinations


def operator_key(name: str) -> str:
    replacements = str.maketrans({"ø": "o", "Ø": "O", "æ": "ae", "Æ": "AE", "ß": "ss"})
    normalized = unicodedata.normalize("NFKD", name.translate(replacements))
    ascii_name = "".join(
        character for character in normalized if not unicodedata.combining(character)
    )
    words = re.findall(r"[a-z0-9]+", ascii_name.lower())
    return "-".join(words)


def format_card_text(card: OperatorCard) -> str:
    primary = "/".join(str(rate) for rate in card.primary_rpms) or "-"
    secondary = "/".join(str(rate) for rate in card.secondary_rpms) or "-"
    semiautomatic = "✓" if card.has_semiautomatic else "-"
    secondary_shotgun = "✓" if card.has_secondary_shotgun else "-"
    return (
        f"{card.name}\n"
        f"{card.speed}速 · 主 {primary} · 副 {secondary}\n"
        f"主狙 {semiautomatic} · 副喷 {secondary_shotgun}"
    )


def write_tier_workbook(
    path: Path,
    cards: Mapping[str, Iterable[OperatorCard]],
    operator_icon_dir: Path,
    gadget_icons: Mapping[str, Path],
    report_sources: ReportSources,
) -> None:
    output_path = Path(path)
    badge_dir = Path(operator_icon_dir)
    normalized_cards = {
        side: sorted(tuple(cards.get(side, ())), key=lambda card: card.source_order)
        for side in SOURCE_SHEETS
    }
    _validate_render_inputs(normalized_cards, badge_dir, gadget_icons)

    workbook = Workbook()
    workbook.remove(workbook.active)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory() as temporary_directory:
        token_dir = Path(temporary_directory)
        token_paths = _make_gadget_tokens(normalized_cards, gadget_icons, token_dir)
        for side in SOURCE_SHEETS:
            sheet = workbook.create_sheet(side + "简图")
            _render_side_sheet(
                sheet,
                side,
                normalized_cards[side],
                badge_dir,
                token_paths,
            )
            _, footer_end = append_source_footer(
                sheet,
                1 + CARDS_PER_ROW * 2,
                report_sources,
            )
            sheet.print_area = "A1:%s%d" % (
                get_column_letter(1 + CARDS_PER_ROW * 2),
                footer_end,
            )
        scores = {
            card.name: int(card.score)
            for side in SOURCE_SHEETS
            for card in normalized_cards[side]
        }
        add_patch_notes_sheet(workbook, scores, report_sources)
        try:
            workbook.save(output_path)
        except OSError as exc:
            raise TierChartError(f"无法保存输出工作簿 {output_path}：{exc}") from exc


def _validate_render_inputs(
    cards: Mapping[str, Tuple[OperatorCard, ...]],
    badge_dir: Path,
    gadget_icons: Mapping[str, Path],
) -> None:
    for side in SOURCE_SHEETS:
        for card in cards[side]:
            if card.side != side:
                raise TierChartError(f"干员阵营与工作表不一致：{card.name}")
            if card.tier not in TIER_ORDER:
                raise TierChartError(f"未知阶层：{card.name} = {card.tier}")
            badge = badge_dir / (operator_key(card.name) + ".png")
            if not badge.is_file() or not _is_valid_image(badge):
                raise TierChartError(f"找不到干员 Badge 或图片损坏：{card.name} ({badge})")
            for gadget in card.gadgets:
                icon = gadget_icons.get(gadget.name)
                if icon is None or not Path(icon).is_file() or not _is_valid_image(Path(icon)):
                    raise TierChartError(f"找不到次要装备图标或图片损坏：{card.name} / {gadget.name}")


def _make_gadget_tokens(
    cards: Mapping[str, Tuple[OperatorCard, ...]],
    gadget_icons: Mapping[str, Path],
    directory: Path,
) -> Dict[Tuple[str, Optional[int]], Path]:
    keys = {
        (gadget.name, gadget.quantity)
        for side in SOURCE_SHEETS
        for card in cards[side]
        for gadget in card.gadgets
    }
    token_paths = {}
    for index, key in enumerate(sorted(keys, key=lambda item: (item[0], item[1] or 0))):
        name, quantity = key
        destination = directory / f"gadget-{index}.png"
        _draw_gadget_token(Path(gadget_icons[name]), quantity, destination)
        token_paths[key] = destination
    return token_paths


def _draw_gadget_token(source: Path, quantity: Optional[int], destination: Path) -> None:
    canvas = PillowImage.new("RGBA", (24, 20), (0, 0, 0, 0))
    with PillowImage.open(source) as image:
        icon = image.convert("RGBA")
        visible_bounds = icon.getchannel("A").getbbox()
        if visible_bounds is None:
            raise TierChartError(f"次要装备图标没有可见像素：{source}")
        icon = icon.crop(visible_bounds)
        scale = min(18 / icon.width, 18 / icon.height)
        icon = icon.resize(
            (max(1, round(icon.width * scale)), max(1, round(icon.height * scale))),
            PillowImage.Resampling.LANCZOS,
        )
        canvas.alpha_composite(icon, ((18 - icon.width) // 2, (20 - icon.height) // 2))
    if quantity is not None:
        draw = ImageDraw.Draw(canvas)
        draw.rounded_rectangle(
            (13, 8, 23, 19),
            radius=2,
            fill=(255, 255, 255, 245),
            outline=(20, 22, 26, 255),
            width=1,
        )
        text = str(quantity)
        font = ImageFont.load_default()
        box = draw.textbbox((0, 0), text, font=font)
        width = box[2] - box[0]
        height = box[3] - box[1]
        draw.text(
            (18 - width / 2, 13.5 - height / 2 - box[1]),
            text,
            font=font,
            fill=(20, 22, 26, 255),
        )
    canvas.save(destination)


def _render_side_sheet(
    sheet,
    side: str,
    cards: Tuple[OperatorCard, ...],
    badge_dir: Path,
    token_paths: Mapping[Tuple[str, Optional[int]], Path],
) -> None:
    last_column = 1 + CARDS_PER_ROW * 2
    last_column_letter = get_column_letter(last_column)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title = sheet.cell(1, 1, f"{side} · Y11S2 SoloQ Tier List")
    title.font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="17191D")
    title.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 5

    sheet.column_dimensions["A"].width = 8
    for slot in range(CARDS_PER_ROW):
        badge_column = 2 + slot * 2
        sheet.column_dimensions[get_column_letter(badge_column)].width = 8
        sheet.column_dimensions[get_column_letter(badge_column + 1)].width = 23

    thin_gray = Side(style="thin", color="CDD1D5")
    card_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    card_fill = PatternFill("solid", fgColor="F4F5F7")
    current_row = 3
    for tier in TIER_ORDER:
        tier_cards = [card for card in cards if card.tier == tier]
        chunks = [tier_cards[index:index + CARDS_PER_ROW] for index in range(0, len(tier_cards), CARDS_PER_ROW)]
        if not chunks:
            chunks = [[]]
        tier_start = current_row
        for chunk in chunks:
            info_row = current_row
            gadget_row = current_row + 1
            sheet.row_dimensions[info_row].height = 52
            sheet.row_dimensions[gadget_row].height = 20
            for slot in range(CARDS_PER_ROW):
                badge_column = 2 + slot * 2
                text_column = badge_column + 1
                for row in (info_row, gadget_row):
                    for column in (badge_column, text_column):
                        cell = sheet.cell(row, column)
                        cell.fill = card_fill
                        cell.border = card_border
                sheet.merge_cells(
                    start_row=gadget_row,
                    start_column=badge_column,
                    end_row=gadget_row,
                    end_column=text_column,
                )
                if slot >= len(chunk):
                    continue
                card = chunk[slot]
                text_cell = sheet.cell(info_row, text_column, format_card_text(card))
                text_cell.font = Font(name="Microsoft YaHei", size=9, color="202327")
                text_cell.alignment = Alignment(vertical="center", wrap_text=True)

                badge = _excel_image(badge_dir / (operator_key(card.name) + ".png"))
                _add_offset_image(sheet, badge, badge_column, info_row, 48, 48, 4, 2)
                for gadget_index, gadget in enumerate(card.gadgets):
                    token = _excel_image(token_paths[(gadget.name, gadget.quantity)])
                    _add_offset_image(
                        sheet,
                        token,
                        badge_column,
                        gadget_row,
                        24,
                        20,
                        4 + gadget_index * 24,
                        0,
                    )
            current_row += 2

        tier_end = current_row - 1
        sheet.merge_cells(start_row=tier_start, start_column=1, end_row=tier_end, end_column=1)
        tier_cell = sheet.cell(tier_start, 1, tier)
        tier_cell.fill = PatternFill("solid", fgColor=TIER_COLORS[tier])
        tier_cell.font = Font(
            name="Arial",
            size=14,
            bold=True,
            color="FFFFFF" if tier not in ("B", "C") else "1B1D20",
        )
        tier_cell.alignment = Alignment(horizontal="center", vertical="center")
        tier_cell.border = card_border

    sheet.freeze_panes = "B3"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 55
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(
        fitToPage=True, autoPageBreaks=False
    )
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.3
    sheet.page_margins.bottom = 0.3
    sheet.print_options.horizontalCentered = True
    sheet.print_area = f"A1:{last_column_letter}{current_row - 1}"


def _add_offset_image(
    sheet,
    image: ExcelImage,
    column: int,
    row: int,
    width: int,
    height: int,
    x_offset: int,
    y_offset: int,
) -> None:
    marker = AnchorMarker(
        col=column - 1,
        colOff=pixels_to_EMU(x_offset),
        row=row - 1,
        rowOff=pixels_to_EMU(y_offset),
    )
    extent = XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height))
    image.width = width
    image.height = height
    image.anchor = OneCellAnchor(_from=marker, ext=extent)
    sheet.add_image(image)


def _excel_image(path: Path) -> ExcelImage:
    return ExcelImage(io.BytesIO(Path(path).read_bytes()))


def main(
    argv: Optional[List[str]] = None,
    *,
    gadget_icon_preparer: Callable[
        [Iterable[GadgetItem], Path], Mapping[str, Path]
    ] = prepare_gadget_icons,
    source_loader: Callable[[Path], ReportSources] = load_report_sources,
) -> int:
    parser = argparse.ArgumentParser(
        description="从 R6 干员统计工作簿生成横向评分阶层简图"
    )
    parser.add_argument("--inputs-dir", type=Path, default=Path("inputs"))
    parser.add_argument("--input", type=Path, default=None)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("~outputs/视频评分简图.xlsx"),
    )
    parser.add_argument(
        "--icons-dir", type=Path, default=None
    )
    parser.add_argument(
        "--gadget-icons-dir", type=Path, default=None
    )
    arguments = parser.parse_args(argv)

    try:
        input_path = arguments.input or Path("~temp") / "r6_operator_stats.xlsx"
        icon_dir = (
            arguments.icons_dir
            or arguments.inputs_dir / "icons" / "operator" / "badge"
        )
        gadget_dir = (
            arguments.gadget_icons_dir
            or arguments.inputs_dir / "icons" / "gadget"
        )
        cards = load_operator_cards(input_path)
        for side in SOURCE_SHEETS:
            for card in cards[side]:
                badge = icon_dir / (operator_key(card.name) + ".png")
                if not badge.is_file() or not _is_valid_image(badge):
                    raise TierChartError(
                        "找不到干员 Badge 或图片损坏：%s (%s)"
                        % (card.name, badge)
                    )
        gadget_items = tuple(
            gadget
            for side in SOURCE_SHEETS
            for card in cards[side]
            for gadget in card.gadgets
        )
        gadget_icons = gadget_icon_preparer(gadget_items, gadget_dir)
        report_sources = source_loader(arguments.inputs_dir)
        write_tier_workbook(
            arguments.output,
            cards,
            icon_dir,
            gadget_icons,
            report_sources,
        )
    except (TierChartError, PatchNotesError, SourceDataError) as exc:
        print(f"错误：{exc}", file=sys.stderr)
        return 1

    print(f"进攻方干员：{len(cards['进攻方'])}")
    print(f"防守方干员：{len(cards['防守方'])}")
    print(f"输出文件：{arguments.output.resolve()}")
    return 0


def _query_huiji_json(parameters: Mapping[str, str]) -> Mapping[str, object]:
    curl_path = shutil.which("curl.exe") or shutil.which("curl")
    if not curl_path:
        raise TierChartError("未找到 curl.exe 或 curl，请先安装并加入 PATH")
    url = API_URL + "?" + urlencode(parameters)
    last_error = "未知请求错误"
    for attempt in range(1, 5):
        try:
            result = subprocess.run(
                [
                    curl_path,
                    "--location",
                    "--silent",
                    "--show-error",
                    "--fail",
                    "--max-time",
                    "30",
                    "--user-agent",
                    "r6-tier-chart/1.0",
                    url,
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="strict",
                check=True,
            )
            if not result.stdout.strip():
                raise ValueError("curl 返回空响应")
            response = json.loads(result.stdout)
            if not isinstance(response, dict):
                raise ValueError("API JSON 顶层不是对象")
            return response
        except (OSError, subprocess.CalledProcessError, UnicodeError, json.JSONDecodeError, ValueError) as exc:
            last_error = str(exc)
            if attempt < 4:
                time.sleep(float(attempt))
    raise TierChartError(f"灰机 Wiki API 经过 4 次尝试仍失败：{last_error}")


def _gadget_filename(name: str) -> str:
    filename = GADGET_FILES[name].split(":", 1)[1]
    filename = re.sub(r"^R6S gp ", "", filename, flags=re.IGNORECASE)
    return filename.replace(" ", "-").lower()


def _is_valid_image(path: Path) -> bool:
    try:
        with PillowImage.open(path) as image:
            image.verify()
        return True
    except (OSError, ValueError):
        return False


def _required_text(value: object, label: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise TierChartError(f"{label}必须是非空文本")
    return value.strip()


def _integer(value: object, message: str) -> int:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise TierChartError(message)
    number = float(value)
    if not number.is_integer():
        raise TierChartError(message)
    return int(number)


def _has_value(value: object) -> bool:
    text = _required_text(value, "武器状态")
    return text != "无"


if __name__ == "__main__":
    raise SystemExit(main())
