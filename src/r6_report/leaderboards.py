"""Build five ordered R6 operator leaderboard workbooks."""

import argparse
import re
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Mapping, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from PIL import Image as PillowImage

from .gadget_slots import GADGET_SLOT_NAMES, arrange_gadgets
from .patch_notes import (
    PatchNotesError,
    add_patch_notes_sheet,
)
from . import report_theme as theme
from .sources import (
    PatchChange,
    ReportSources,
    SourceDataError,
    load_report_sources,
)
from .tier_chart import (
    CARDS_PER_ROW,
    GadgetItem,
    OperatorCard,
    SOURCE_SHEETS,
    TierChartError,
    _add_offset_image,
    _excel_image,
    _gadget_filename,
    _is_valid_image,
    _validate_render_inputs,
    load_operator_cards,
    operator_key,
    prepare_gadget_icons,
)
from .tiers import TIER_COLORS
from .workbook_sources import append_source_footer


DIMENSION_ORDER: Tuple[str, ...] = (
    "video",
    "primary_rpm",
    "speed",
    "rare",
    "gadget",
)
VIDEO_BANDS: Tuple[str, ...] = ("S", "A", "B", "C", "D", "F")
PRIMARY_RPM_BANDS: Tuple[str, ...] = ("Ⅰ", "Ⅱ", "Ⅲ", "Ⅳ")
SPEED_BANDS: Tuple[str, ...] = ("3速", "2速", "1速")
RARE_BANDS: Tuple[str, ...] = ("副喷", "主狙", "副自", "都无")
ATTACK_GADGET_BANDS: Tuple[str, ...] = (
    "手雷",
    "眩晕手榴弹",
    "硬突破炸药",
    "这些都无",
)
DEFENSE_GADGET_BANDS: Tuple[str, ...] = (
    "遥控炸药",
    "机动护盾",
    "冲击手榴弹",
    "这些都无",
)
ATTACK_GADGET_NAMES = {
    "手雷": "破片手榴弹",
    "眩晕手榴弹": "闪光弹",
    "硬突破炸药": "硬突破炸药",
}
DEFENSE_GADGET_NAMES = {
    "遥控炸药": "遥控炸药",
    "机动护盾": "机动护盾",
    "冲击手榴弹": "冲击手榴弹",
}
PATCH_DIRECTIONS = {"增强", "削弱", "混合"}
SECONDARY_GADGET_NAMES = frozenset(
    name
    for slot_names in GADGET_SLOT_NAMES.values()
    for name in slot_names
    if name is not None
)
_GADGET_EXPRESSION = "(?:%s)" % "|".join(
    re.escape(name)
    for name in sorted(SECONDARY_GADGET_NAMES, key=len, reverse=True)
)
_LOADOUT_CHANGE_PATTERNS = tuple(
    re.compile(pattern)
    for pattern in (
        r"新增\s*" + _GADGET_EXPRESSION,
        r"移除\s*" + _GADGET_EXPRESSION,
        _GADGET_EXPRESSION
        + r"\s*(?:被)?替换为\s*"
        + _GADGET_EXPRESSION,
        r"改为\s*配备\s*" + _GADGET_EXPRESSION,
        r"配备\s*" + _GADGET_EXPRESSION,
    )
)
_PATCH_DETAIL_SEPARATOR = re.compile(r"[。；\r\n]+")


class LeaderboardError(Exception):
    """Raised when leaderboard data or configuration is invalid."""


@dataclass(frozen=True)
class LeaderboardSpec:
    key: str
    filename: str
    sheet_suffix: str
    title: str
    repeated: bool = False


LEADERBOARD_SPECS: Mapping[str, LeaderboardSpec] = {
    "video": LeaderboardSpec(
        "video",
        "视频评分榜.xlsx",
        "视频Tier榜",
        "视频 Tier 榜",
    ),
    "primary_rpm": LeaderboardSpec(
        "primary_rpm",
        "主武器射速榜.xlsx",
        "主武器射速榜",
        "主武器射速榜",
    ),
    "speed": LeaderboardSpec(
        "speed",
        "速度榜.xlsx",
        "速度榜",
        "速度榜",
    ),
    "rare": LeaderboardSpec(
        "rare",
        "稀有枪械榜.xlsx",
        "稀有枪械榜",
        "稀有枪械榜",
        repeated=True,
    ),
    "gadget": LeaderboardSpec(
        "gadget",
        "次要装备榜.xlsx",
        "次要装备榜",
        "次要装备榜",
        repeated=True,
    ),
}
EXPECTED_OUTPUTS: Tuple[str, ...] = tuple(
    spec.filename for spec in LEADERBOARD_SPECS.values()
)


def patch_direction_marker(directions: Iterable[str]) -> str:
    values = set(directions)
    unknown = values - PATCH_DIRECTIONS
    if unknown:
        raise LeaderboardError(
            "未知补丁方向：%s" % ", ".join(sorted(unknown))
        )
    if not values:
        return ""
    if values == {"增强"}:
        return "+"
    if values == {"削弱"}:
        return "-"
    return "~"


def counts_as_operator_adjustment(change: PatchChange) -> bool:
    """判断补丁条目是否应计入干员增强或削弱标记。"""
    detail = change.detail.strip()
    if not any(name in detail for name in SECONDARY_GADGET_NAMES):
        return True
    if any(pattern.search(detail) for pattern in _LOADOUT_CHANGE_PATTERNS):
        return True
    first_clause = next(
        (
            clause.strip()
            for clause in _PATCH_DETAIL_SEPARATOR.split(detail)
            if clause.strip()
        ),
        "",
    )
    return not any(
        name in first_clause
        for name in SECONDARY_GADGET_NAMES
    )


def patch_markers(sources: ReportSources) -> Mapping[str, str]:
    directions: Dict[str, List[str]] = {}
    for patch in sources.patches:
        for change in patch.changes:
            if not counts_as_operator_adjustment(change):
                continue
            directions.setdefault(change.subject, []).append(change.direction)
    return {
        subject: patch_direction_marker(values)
        for subject, values in directions.items()
    }


def band_order(dimension: str, side: str) -> Tuple[str, ...]:
    _validate_dimension_and_side(dimension, side)
    if dimension == "video":
        return VIDEO_BANDS
    if dimension == "primary_rpm":
        return PRIMARY_RPM_BANDS
    if dimension == "speed":
        return SPEED_BANDS
    if dimension == "rare":
        return RARE_BANDS
    if side == "进攻方":
        return ATTACK_GADGET_BANDS
    return DEFENSE_GADGET_BANDS


def bands_for_card(
    card: OperatorCard,
    dimension: str,
    side: str,
) -> Tuple[str, ...]:
    _validate_dimension_and_side(dimension, side)
    if card.side != side:
        raise LeaderboardError("干员阵营与榜单不一致：%s" % card.name)

    if dimension == "video":
        if card.tier not in VIDEO_BANDS:
            raise LeaderboardError("未知视频 Tier：%s = %s" % (card.name, card.tier))
        return (card.tier,)

    if dimension == "primary_rpm":
        rate = max(card.primary_rpms) if card.primary_rpms else None
        if rate is None or rate < 700:
            return ("Ⅳ",)
        if rate >= 860:
            return ("Ⅰ",)
        if rate >= 780:
            return ("Ⅱ",)
        return ("Ⅲ",)

    if dimension == "speed":
        if card.speed not in (1, 2, 3):
            raise LeaderboardError("未知速度：%s = %s" % (card.name, card.speed))
        return ("%d速" % card.speed,)

    if dimension == "rare":
        memberships = []
        if card.has_secondary_shotgun:
            memberships.append("副喷")
        if card.has_semiautomatic:
            memberships.append("主狙")
        if card.secondary_rpms:
            memberships.append("副自")
        return tuple(memberships) or ("都无",)

    gadget_names = {gadget.name for gadget in card.gadgets}
    category_names = (
        ATTACK_GADGET_NAMES
        if side == "进攻方"
        else DEFENSE_GADGET_NAMES
    )
    memberships = [
        category
        for category, gadget_name in category_names.items()
        if gadget_name in gadget_names
    ]
    return tuple(memberships) or ("这些都无",)


def best_dimension_rank(
    card: OperatorCard,
    dimension: str,
    side: str,
) -> int:
    order = band_order(dimension, side)
    memberships = bands_for_card(card, dimension, side)
    return min(order.index(item) for item in memberships)


def sort_cards_for_band(
    cards: Iterable[OperatorCard],
    current_dimension: str,
    side: str,
) -> List[OperatorCard]:
    _validate_dimension_and_side(current_dimension, side)
    dimensions = tuple(
        dimension
        for dimension in DIMENSION_ORDER
        if dimension != current_dimension
    )

    def sort_key(card: OperatorCard):
        return tuple(
            best_dimension_rank(card, dimension, side)
            for dimension in dimensions
        ) + (card.source_order,)

    return sorted(cards, key=sort_key)


def group_cards(
    cards: Iterable[OperatorCard],
    dimension: str,
    side: str,
) -> Mapping[str, Tuple[OperatorCard, ...]]:
    order = band_order(dimension, side)
    spec = LEADERBOARD_SPECS[dimension]
    groups: Dict[str, List[OperatorCard]] = {band: [] for band in order}
    source_cards = tuple(cards)
    names = [card.name for card in source_cards]
    if len(names) != len(set(names)):
        raise LeaderboardError("输入存在重复干员：%s" % side)

    for card in source_cards:
        memberships = bands_for_card(card, dimension, side)
        if not spec.repeated and len(memberships) != 1:
            raise LeaderboardError("非重复榜单出现多级干员：%s" % card.name)
        for membership in memberships:
            groups[membership].append(card)

    covered = {
        card.name
        for group in groups.values()
        for card in group
    }
    if covered != set(names):
        raise LeaderboardError("榜单未覆盖全部干员：%s" % side)

    return {
        band: tuple(sort_cards_for_band(groups[band], dimension, side))
        for band in order
    }


def write_leaderboard_workbook(
    path: Path,
    spec: LeaderboardSpec,
    cards: Mapping[str, Iterable[OperatorCard]],
    operator_icon_dir: Path,
    gadget_icons: Mapping[str, Path],
    report_sources: ReportSources,
) -> None:
    output_path = Path(path)
    badge_dir = Path(operator_icon_dir)
    normalized_cards = {
        side: tuple(
            sorted(
                tuple(cards.get(side, ())),
                key=lambda card: card.source_order,
            )
        )
        for side in SOURCE_SHEETS
    }
    _validate_render_inputs(normalized_cards, badge_dir, gadget_icons)

    workbook = Workbook()
    workbook.remove(workbook.active)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory() as temporary_directory:
        token_paths = _make_gadget_tokens(
            normalized_cards,
            gadget_icons,
            Path(temporary_directory),
        )
        markers = patch_markers(report_sources)
        for side in SOURCE_SHEETS:
            sheet = workbook.create_sheet(side + spec.sheet_suffix)
            groups = group_cards(normalized_cards[side], spec.key, side)
            _render_side_sheet(
                sheet,
                side,
                spec,
                groups,
                badge_dir,
                token_paths,
                markers,
            )
            _, footer_end = append_source_footer(
                sheet,
                1 + CARDS_PER_ROW * 6,
                report_sources,
            )
            sheet.print_area = "A1:%s%d" % (
                get_column_letter(1 + CARDS_PER_ROW * 6),
                footer_end,
            )

        scores = {
            card.name: int(card.score)
            for side in SOURCE_SHEETS
            for card in normalized_cards[side]
        }
        add_patch_notes_sheet(workbook, scores, report_sources)
        try:
            workbook.save(output_path)
        except OSError as exc:
            raise LeaderboardError(
                "无法保存输出工作簿 %s：%s" % (output_path, exc)
            ) from exc


def write_all_leaderboards(
    output_dir: Path,
    cards: Mapping[str, Iterable[OperatorCard]],
    operator_icon_dir: Path,
    gadget_icons: Mapping[str, Path],
    report_sources: ReportSources,
) -> Tuple[Path, ...]:
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    paths = []
    for spec in LEADERBOARD_SPECS.values():
        path = destination / spec.filename
        write_leaderboard_workbook(
            path,
            spec,
            cards,
            operator_icon_dir,
            gadget_icons,
            report_sources,
        )
        paths.append(path)
        pdf_path = path.with_suffix(".pdf")
        from .pdf_leaderboards import (
            write_leaderboard_pdf,
            write_pdf_pages_as_png,
        )
        write_leaderboard_pdf(
            pdf_path,
            spec,
            cards,
            operator_icon_dir,
            gadget_icons,
            report_sources,
        )
        paths.append(pdf_path)
        paths.extend(
            write_pdf_pages_as_png(
                pdf_path,
                destination / "图片版" / pdf_path.stem,
            )
        )
    return tuple(paths)


def _make_gadget_tokens(
    cards: Mapping[str, Tuple[OperatorCard, ...]],
    gadget_icons: Mapping[str, Path],
    directory: Path,
) -> Mapping[Tuple[str, Optional[int]], Path]:
    keys = set()
    for side in SOURCE_SHEETS:
        for card in cards[side]:
            for gadget in card.gadgets:
                keys.add((gadget.name, gadget.quantity))

    paths = {}
    for index, key in enumerate(
        sorted(keys, key=lambda item: (item[0], item[1] or 0))
    ):
        name, quantity = key
        source = gadget_icons.get(name)
        if source is None:
            raise LeaderboardError("找不到次要装备图标：%s" % name)
        destination = directory / ("gadget-%d.png" % index)
        draw_gadget_token(
            Path(source),
            quantity,
            destination,
        )
        paths[key] = destination
    return paths


def draw_gadget_token(
    source: Path,
    quantity: Optional[int],
    destination: Path,
) -> None:
    canvas_width, canvas_height = theme.XLSX_GADGET_TOKEN_PX
    canvas = PillowImage.new(
        "RGBA",
        (canvas_width, canvas_height),
        (0, 0, 0, 0),
    )

    with PillowImage.open(source) as image:
        icon = image.convert("RGBA")
        visible_bounds = icon.getchannel("A").getbbox()
        if visible_bounds is None:
            raise LeaderboardError("次要装备图标没有可见像素：%s" % source)
        icon = icon.crop(visible_bounds)
        scale = min(
            theme.XLSX_GADGET_ICON_BOX_PX / icon.width,
            theme.XLSX_GADGET_ICON_BOX_PX / icon.height,
        )
        icon = icon.resize(
            (
                max(1, round(icon.width * scale)),
                max(1, round(icon.height * scale)),
            ),
            PillowImage.Resampling.LANCZOS,
        )
        canvas.alpha_composite(
            icon,
            (
                (canvas_width - icon.width) // 2,
                (canvas_height - icon.height) // 2,
            ),
        )
    canvas.save(destination)


def _render_side_sheet(
    sheet,
    side: str,
    spec: LeaderboardSpec,
    groups: Mapping[str, Tuple[OperatorCard, ...]],
    badge_dir: Path,
    token_paths: Mapping[Tuple[str, Optional[int]], Path],
    markers: Mapping[str, str],
) -> None:
    columns_per_card = 6
    last_column = 1 + CARDS_PER_ROW * columns_per_card
    last_column_letter = get_column_letter(last_column)
    sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_column,
    )
    title = sheet.cell(1, 1, "%s · %s" % (side, spec.title))
    title.font = Font(
        name=theme.FONT_FAMILY,
        size=theme.XLSX_FONT_SIZES["title"],
        bold=True,
        color=theme.COLOURS["white"],
    )
    title.fill = PatternFill("solid", fgColor=theme.COLOURS["title_fill"])
    title.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 5

    sheet.column_dimensions["A"].width = 9
    for slot in range(CARDS_PER_ROW):
        badge_column = 2 + slot * columns_per_card
        sheet.column_dimensions[get_column_letter(badge_column)].width = 6.5
        for column in range(badge_column + 1, badge_column + 5):
            sheet.column_dimensions[get_column_letter(column)].width = 6.8
        sheet.column_dimensions[
            get_column_letter(badge_column + 5)
        ].width = 2.5

    thin_gray = Side(style="thin", color=theme.COLOURS["border"])
    card_border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )
    card_fill = PatternFill("solid", fgColor=theme.COLOURS["card_fill"])
    missing_fill = PatternFill("solid", fgColor=theme.MISSING_FILL)
    normal_font = Font(
        name=theme.FONT_FAMILY,
        size=theme.XLSX_FONT_SIZES["body"],
        color=theme.COLOURS["text"],
    )
    current_row = 3

    for band in band_order(spec.key, side):
        band_cards = groups[band]
        chunks = [
            band_cards[index:index + CARDS_PER_ROW]
            for index in range(0, len(band_cards), CARDS_PER_ROW)
        ] or [()]
        band_start_row = current_row

        for chunk in chunks:
            name_row = current_row
            feature_row = current_row + 1
            rpm_row = current_row + 2
            gadget_row = current_row + 3
            sheet.row_dimensions[name_row].height = 18
            sheet.row_dimensions[
                feature_row
            ].height = theme.XLSX_CARD_BODY_ROW_PT
            sheet.row_dimensions[
                rpm_row
            ].height = theme.XLSX_CARD_BODY_ROW_PT
            sheet.row_dimensions[gadget_row].height = (
                2 * theme.XLSX_CARD_BODY_ROW_PT
            )

            for slot in range(CARDS_PER_ROW):
                badge_column = 2 + slot * columns_per_card
                first_info = badge_column + 1
                second_info = badge_column + 2
                third_info = badge_column + 3
                fourth_info = badge_column + 4
                for row in (
                    name_row,
                    feature_row,
                    rpm_row,
                    gadget_row,
                ):
                    for column in range(badge_column, fourth_info + 1):
                        cell = sheet.cell(row, column)
                        cell.fill = card_fill
                        cell.border = card_border

                sheet.merge_cells(
                    start_row=name_row,
                    start_column=badge_column,
                    end_row=rpm_row,
                    end_column=badge_column,
                )
                sheet.merge_cells(
                    start_row=name_row,
                    start_column=first_info,
                    end_row=name_row,
                    end_column=second_info,
                )
                for row in (feature_row, rpm_row):
                    sheet.merge_cells(
                        start_row=row,
                        start_column=first_info,
                        end_row=row,
                        end_column=second_info,
                    )
                    sheet.merge_cells(
                        start_row=row,
                        start_column=third_info,
                        end_row=row,
                        end_column=fourth_info,
                    )
                sheet.merge_cells(
                    start_row=gadget_row,
                    start_column=badge_column,
                    end_row=gadget_row,
                    end_column=fourth_info,
                )
                if slot >= len(chunk):
                    continue

                card = chunk[slot]
                name_cell = sheet.cell(name_row, first_info, card.name)
                tier_cell = sheet.cell(
                    name_row,
                    third_info,
                    card.tier + markers.get(card.name, ""),
                )
                speed_cell = sheet.cell(
                    name_row,
                    fourth_info,
                    "%d速" % card.speed,
                )
                secondary_shotgun_cell = sheet.cell(
                    feature_row,
                    first_info,
                    theme.feature_text("副喷", card.has_secondary_shotgun),
                )
                primary_sniper_cell = sheet.cell(
                    feature_row,
                    third_info,
                    theme.feature_text("主狙", card.has_semiautomatic),
                )
                secondary_rpm_cell = sheet.cell(
                    rpm_row,
                    first_info,
                    theme.rpm_text("副", card.secondary_rpms),
                )
                primary_rpm_cell = sheet.cell(
                    rpm_row,
                    third_info,
                    theme.rpm_text("主", card.primary_rpms),
                )

                for cell in (
                    name_cell,
                    speed_cell,
                    secondary_shotgun_cell,
                    primary_sniper_cell,
                    secondary_rpm_cell,
                    primary_rpm_cell,
                ):
                    cell.font = normal_font
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        shrink_to_fit=True,
                    )
                    if theme.is_missing_field(cell.value):
                        cell.fill = missing_fill
                name_cell.font = Font(
                    name=theme.FONT_FAMILY,
                    size=theme.XLSX_FONT_SIZES["name"],
                    color=theme.COLOURS["text"],
                )
                tier_cell.font = Font(
                    name="Arial",
                    size=theme.XLSX_FONT_SIZES["body"],
                    bold=True,
                    color=(
                        theme.COLOURS["white"]
                        if card.tier in ("S", "A", "D", "F")
                        else theme.COLOURS["text_strong"]
                    ),
                )
                tier_cell.fill = PatternFill(
                    "solid",
                    fgColor=TIER_COLORS[card.tier],
                )
                tier_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )
                speed_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

                badge = _excel_image(
                    badge_dir / (operator_key(card.name) + ".png")
                )
                _add_offset_image(
                    sheet,
                    badge,
                    badge_column,
                    name_row,
                    48,
                    48,
                    4,
                    2,
                )
                for gadget_slot, gadget in enumerate(
                    arrange_gadgets(card.side, card.gadgets)
                ):
                    if gadget is None:
                        continue
                    gadget_column = (
                        gadget_slot % theme.GADGETS_PER_LINE
                    )
                    gadget_line = gadget_slot // theme.GADGETS_PER_LINE
                    token = _excel_image(
                        token_paths[(gadget.name, gadget.quantity)]
                    )
                    _add_offset_image(
                        sheet,
                        token,
                        first_info + gadget_column,
                        gadget_row,
                        theme.XLSX_GADGET_TOKEN_PX[0],
                        theme.XLSX_GADGET_TOKEN_PX[1],
                        theme.XLSX_GADGET_COLUMN_OFFSET_PX,
                        gadget_line
                        * theme.XLSX_GADGET_TOKEN_PX[1],
                    )
            current_row += 4

        band_end_row = current_row - 1
        sheet.merge_cells(
            start_row=band_start_row,
            start_column=1,
            end_row=band_end_row,
            end_column=1,
        )
        band_cell = sheet.cell(band_start_row, 1, band)
        band_color = _band_color(spec.key, band, side)
        band_cell.fill = PatternFill("solid", fgColor=band_color)
        band_cell.font = Font(
            name=theme.FONT_FAMILY,
            size=theme.XLSX_FONT_SIZES["band"],
            bold=True,
            color=_contrast_text_color(band_color),
        )
        band_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        band_cell.border = card_border

    sheet.freeze_panes = "B3"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 48
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(
        fitToPage=True,
        autoPageBreaks=False,
    )
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.3
    sheet.page_margins.bottom = 0.3
    sheet.print_options.horizontalCentered = True
    sheet.print_area = "A1:%s%d" % (
        last_column_letter,
        current_row - 1,
    )

def _band_color(dimension: str, band: str, side: str) -> str:
    if dimension == "video":
        return TIER_COLORS[band]
    palettes = {
        "primary_rpm": {
            "Ⅰ": "E74C3C",
            "Ⅱ": "F39C12",
            "Ⅲ": "F1C40F",
            "Ⅳ": "7F8C8D",
        },
        "speed": {
            "3速": "E74C3C",
            "2速": "F39C12",
            "1速": "7F8C8D",
        },
        "rare": {
            "副喷": "2A6F6B",
            "主狙": "4472C4",
            "副自": "D4A72C",
            "都无": "7F8C8D",
        },
        "gadget": {
            band_order("gadget", side)[0]: "E74C3C",
            band_order("gadget", side)[1]: "F39C12",
            band_order("gadget", side)[2]: "F1C40F",
            band_order("gadget", side)[3]: "7F8C8D",
        },
    }
    return palettes[dimension][band]


def _contrast_text_color(color: str) -> str:
    red, green, blue = bytes.fromhex(color)
    luminance = 0.2126 * red + 0.7152 * green + 0.0722 * blue
    return "1B1D20" if luminance > 155 else "FFFFFF"


def _validate_dimension_and_side(dimension: str, side: str) -> None:
    if dimension not in LEADERBOARD_SPECS:
        raise LeaderboardError("未知榜单维度：%s" % dimension)
    if side not in SOURCE_SHEETS:
        raise LeaderboardError("未知阵营：%s" % side)


def load_gadget_icons(
    items: Iterable[GadgetItem],
    directory: Path,
) -> Mapping[str, Path]:
    """Load already collected gadget icons without network access."""
    root = Path(directory)
    paths = {}
    for item in items:
        if item.name in paths:
            continue
        path = root / _gadget_filename(item.name)
        if not path.is_file() or not _is_valid_image(path):
            raise LeaderboardError(
                "找不到已采集的次要装备图标：%s (%s)"
                % (item.name, path)
            )
        paths[item.name] = path
    return paths


def main(
    argv: Optional[List[str]] = None,
    *,
    card_loader: Callable[
        [Path], Mapping[str, List[OperatorCard]]
    ] = load_operator_cards,
    gadget_icon_preparer: Callable[
        [Iterable[GadgetItem], Path], Mapping[str, Path]
    ] = None,
    source_loader: Callable[[Path], ReportSources] = load_report_sources,
) -> int:
    parser = argparse.ArgumentParser(
        description="从 R6 干员统计工作簿生成五个榜单工作簿"
    )
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=Path("data"),
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output"),
    )
    parser.add_argument(
        "--icons-dir",
        type=Path,
        default=None,
    )
    parser.add_argument(
        "--gadget-icons-dir",
        type=Path,
        default=None,
    )
    arguments = parser.parse_args(argv)

    try:
        input_path = (
            arguments.input
            or arguments.data_dir / "r6_operator_stats.xlsx"
        )
        operator_icon_dir = (
            arguments.icons_dir
            or arguments.data_dir / "icons" / "operator" / "badge"
        )
        gadget_icon_dir = (
            arguments.gadget_icons_dir
            or arguments.data_dir / "icons" / "gadget"
        )
        if not input_path.is_file():
            raise LeaderboardError(
                "找不到输入文件：%s" % input_path
            )
        cards = card_loader(input_path)
        gadget_items = tuple(
            gadget
            for side in SOURCE_SHEETS
            for card in cards[side]
            for gadget in card.gadgets
        )
        gadget_icons = (
            gadget_icon_preparer(gadget_items, gadget_icon_dir)
            if gadget_icon_preparer is not None
            else load_gadget_icons(gadget_items, gadget_icon_dir)
        )
        report_sources = source_loader(arguments.data_dir)
        output_paths = write_all_leaderboards(
            arguments.output_dir,
            cards,
            operator_icon_dir,
            gadget_icons,
            report_sources,
        )
    except (
        LeaderboardError,
        TierChartError,
        PatchNotesError,
        SourceDataError,
    ) as exc:
        print("错误：%s" % exc, file=sys.stderr)
        return 1

    print("进攻方干员：%d" % len(cards["进攻方"]))
    print("防守方干员：%d" % len(cards["防守方"]))
    for path in output_paths:
        print("输出文件：%s" % path.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
