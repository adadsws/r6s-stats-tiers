"""Shared source metadata footer rendering for generated workbooks."""

from typing import Tuple

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .sources import ReportSources
from . import report_theme as theme


class WorkbookSourceError(ValueError):
    """Raised when a source footer cannot be rendered safely."""


def append_source_footer(
    sheet: Worksheet,
    last_column: int,
    sources: ReportSources,
) -> Tuple[int, int]:
    """Append three merged source rows and return their inclusive row range."""
    if last_column < 1:
        raise WorkbookSourceError("last_column must be positive")

    first_row = sheet.max_row + 2
    rating = sources.rating
    wiki = sources.wiki
    if sources.patches:
        interval = "%s（%s）→ %s（%s）" % (
            sources.patches[0].patch,
            sources.patches[0].released.isoformat(),
            sources.patches[-1].patch,
            sources.patches[-1].released.isoformat(),
        )
    else:
        interval = "评分覆盖日期之后无补丁"

    rows = (
        (
            "评分来源：%s《%s》｜%s｜覆盖至 %s（%s）｜发布 %s｜采集 %s"
            % (
                rating.creator,
                rating.title,
                rating.season,
                rating.covered_patch,
                rating.covered_through.isoformat(),
                rating.published.isoformat(),
                rating.captured_at.isoformat(),
            ),
            rating.url,
        ),
        (
            "游戏数据：灰机 Wiki｜%s %s｜%s｜抓取 %s｜"
            "除评分外，其他信息均为该时间点的最新数据"
            % (
                wiki.season,
                wiki.season_name,
                wiki.patch,
                wiki.fetched_at.isoformat(),
            ),
            wiki.sources["operator"],
        ),
        (
            "补丁区间：评分覆盖至 %s；Wiki 数据截至 %s；%s"
            % (
                rating.covered_through.isoformat(),
                wiki.fetched_at.date().isoformat(),
                interval,
            ),
            sources.patch_index_url,
        ),
    )

    fill = PatternFill("solid", fgColor=theme.MISSING_FILL)
    font = Font(
        name=theme.FONT_FAMILY,
        size=theme.XLSX_FONT_SIZES["source"],
        italic=True,
        color=theme.COLOURS["text_muted"],
    )
    for offset, (text, url) in enumerate(rows):
        row = first_row + offset
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=last_column,
        )
        cell = sheet.cell(row, 1, text)
        cell.hyperlink = url
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[row].height = 27
    return first_row, first_row + len(rows) - 1
