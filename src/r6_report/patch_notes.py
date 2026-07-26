"""Render source-linked post-video patch notes from collected snapshots."""

from typing import Mapping

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

from .sources import PatchRecord, ReportSources
from .tiers import SCORE_TO_DISPLAY_TIER
from .workbook_sources import append_source_footer


PATCH_SHEET_TITLE = "补丁说明"
DIRECTION_COLOURS = {
    "增强": "548235",
    "削弱": "C00000",
    "混合": "BF9000",
}


class PatchNotesError(ValueError):
    """Raised when patch metadata cannot be rendered safely."""


def _score_text(name: str, scores: Mapping[str, int]) -> str:
    if name not in scores:
        raise PatchNotesError("missing video score for patch subject: %s" % name)
    score = scores[name]
    if isinstance(score, bool):
        raise PatchNotesError(
            "unknown video score for patch subject: %s = %s" % (name, score)
        )
    tier = SCORE_TO_DISPLAY_TIER.get(score)
    if tier is None:
        raise PatchNotesError(
            "unknown video score for patch subject: %s = %s" % (name, score)
        )
    return "%s / %d" % (tier, score)


def _validate_metadata(
    scores: Mapping[str, int],
    sources: ReportSources,
) -> None:
    seen = set()
    for patch in sources.patches:
        for url in (patch.wiki_url, patch.official_url):
            if not url.startswith("https://"):
                raise PatchNotesError(
                    "patch source must use HTTPS: %s" % url
                )
        for change in patch.changes:
            if change.direction not in DIRECTION_COLOURS:
                raise PatchNotesError(
                    "unknown patch direction: %s" % change.direction
                )
            key = (patch.patch, change.subject, change.detail)
            if key in seen:
                raise PatchNotesError(
                    "duplicate patch change: %s / %s"
                    % (patch.patch, change.subject)
                )
            seen.add(key)
            _score_text(change.subject, scores)


def _style_patch_header(sheet: Worksheet, row: int, patch: PatchRecord) -> None:
    sheet.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=6
    )
    cell = sheet.cell(
        row,
        1,
        "%s · %s · %s"
        % (patch.patch, patch.released.isoformat(), patch.season_name),
    )
    cell.fill = PatternFill("solid", fgColor="5B9BD5")
    cell.font = Font(
        name="Microsoft YaHei", size=11, bold=True, color="FFFFFF"
    )
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 25


def _add_source_row(
    sheet: Worksheet,
    row: int,
    patch: PatchRecord,
) -> None:
    sheet.cell(row, 1, "灰机 Wiki")
    sheet.cell(row, 1).font = Font(
        name="Microsoft YaHei", size=8, bold=True
    )
    wiki = sheet.cell(row, 2, patch.wiki_url)
    wiki.hyperlink = patch.wiki_url
    wiki.style = "Hyperlink"
    sheet.cell(row, 4, "Ubisoft")
    sheet.cell(row, 4).font = Font(
        name="Microsoft YaHei", size=8, bold=True
    )
    official = sheet.cell(row, 5, patch.official_url)
    official.hyperlink = patch.official_url
    official.style = "Hyperlink"
    sheet.merge_cells(
        start_row=row, start_column=2, end_row=row, end_column=3
    )
    sheet.merge_cells(
        start_row=row, start_column=5, end_row=row, end_column=6
    )
    for column in range(1, 7):
        sheet.cell(row, column).alignment = Alignment(
            vertical="center", wrap_text=True
        )
    sheet.row_dimensions[row].height = 28


def _add_change_header(sheet: Worksheet, row: int) -> None:
    headers = ("方向", "补丁", "日期", "干员/对象", "视频评分", "更新内容")
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row, column, header)
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(
            name="Microsoft YaHei", size=9, bold=True, color="FFFFFF"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[row].height = 22


def _add_patch_changes(
    sheet: Worksheet,
    row: int,
    patch: PatchRecord,
    scores: Mapping[str, int],
) -> int:
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if not patch.changes:
        sheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=6
        )
        cell = sheet.cell(row, 1, "无影响本报告字段的变更")
        cell.font = Font(
            name="Microsoft YaHei", size=9, italic=True, color="595959"
        )
        cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[row].height = 24
        return row + 1

    for change in patch.changes:
        values = (
            change.direction,
            patch.patch,
            patch.released.isoformat(),
            change.subject,
            _score_text(change.subject, scores),
            change.detail,
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column, value)
            cell.font = Font(
                name="Microsoft YaHei",
                size=9,
                bold=column == 1,
                color="FFFFFF" if column == 1 else "1F1F1F",
            )
            cell.alignment = Alignment(
                horizontal="center" if column < 6 else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border
            if column == 1:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=DIRECTION_COLOURS[change.direction],
                )
        sheet.row_dimensions[row].height = 32
        row += 1
    return row


def add_patch_notes_sheet(
    workbook: Workbook,
    scores_by_name: Mapping[str, int],
    sources: ReportSources,
) -> Worksheet:
    """Append patch notes collected between rating and Wiki snapshot dates."""
    if PATCH_SHEET_TITLE in workbook.sheetnames:
        raise PatchNotesError(
            "patch sheet already exists: %s" % PATCH_SHEET_TITLE
        )
    _validate_metadata(scores_by_name, sources)

    sheet = workbook.create_sheet(PATCH_SHEET_TITLE)
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "%s 视频评分后续补丁说明" % sources.rating.season
    sheet["A1"].font = Font(
        name="Microsoft YaHei", size=18, bold=True, color="FFFFFF"
    )
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34

    sheet.merge_cells("A2:F2")
    sheet["A2"] = (
        "除 Athieno %s 视频评分外，干员、速度、武器、装备、图标与补丁信息"
        "均按灰机 Wiki 抓取时间更新。" % sources.rating.season
    )
    sheet["A2"].font = Font(
        name="Microsoft YaHei", size=10, bold=True, color="1F1F1F"
    )
    sheet["A2"].fill = PatternFill("solid", fgColor="D9EAF7")
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 31

    sheet.merge_cells("A3:F3")
    sheet["A3"] = (
        "补丁范围：%s 之后至 %s；以下内容不重新计算视频评分。"
        % (
            sources.rating.covered_through.isoformat(),
            sources.wiki.fetched_at.date().isoformat(),
        )
    )
    sheet["A3"].font = Font(
        name="Microsoft YaHei", size=9, italic=True, color="595959"
    )
    sheet["A3"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[3].height = 27

    row = 5
    for patch in sources.patches:
        _style_patch_header(sheet, row, patch)
        row += 1
        _add_source_row(sheet, row, patch)
        row += 1
        _add_change_header(sheet, row)
        row += 1
        row = _add_patch_changes(sheet, row, patch, scores_by_name)
        row += 1
    if not sources.patches:
        sheet.merge_cells("A5:F5")
        sheet["A5"] = "评分覆盖日期之后没有需要列出的补丁。"
        row = 6

    for column, width in enumerate((16, 14, 14, 22, 14, 70), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    _, footer_end = append_source_footer(sheet, 6, sources)
    sheet.print_area = "A1:F%d" % footer_end
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(
        fitToPage=True, autoPageBreaks=False
    )
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35
    return sheet
