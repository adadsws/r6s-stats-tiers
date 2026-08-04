"""Build a Rainbow Six operator weapon, gadget, icon, and rating workbook."""

import argparse
import json
import math
import re
import shutil
import subprocess
import sys
import tempfile
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image as PillowImage
from PIL import ImageChops

from .patch_notes import (
    add_patch_notes_sheet,
)
from .sources import ReportSources, SourceDataError, load_report_sources
from .tiers import TIER_COLORS, display_tier
from .workbook_sources import append_source_footer


API_URL = "https://r6s.huijiwiki.com/api.php"
ATTACKERS = "进攻方"
DEFENDERS = "防守方"
SIDES = (ATTACKERS, DEFENDERS)
HEADERS = (
    "图标",
    "干员",
    "速度",
    "主手自动枪械（射速，发/分钟）",
    "副手自动枪械（射速，发/分钟）",
    "主狙",
    "副手霰弹",
    "次要装备",
    "Athieno评分",
)
OPERATOR_FIELDS = ("id", "name", "camp", "speed", "_index", "props")
WEAPON_FIELDS = ("id", "zh_model", "firerate", "projectile", "index", "type", "equipment")
CONFIG_FIELDS = ("id", "user")
DEFAULT_INPUTS_DIR = Path("inputs")
WEAPON_NAME_ALIASES = {
    "P10 RONI转换套件衍生型": "P10 RONI",
}


class R6StatsError(Exception):
    """Base error for expected data collection and processing failures."""


class FetchError(R6StatsError):
    """Raised when a source page cannot be fetched reliably."""


class DataFormatError(R6StatsError):
    """Raised when a source page or its records do not have the expected shape."""


@dataclass(frozen=True)
class WeaponRate:
    name: str
    firerate: float
    order: float
    weapon_id: str = ""


@dataclass(frozen=True)
class OperatorRow:
    name: str
    speed: float
    order: float
    primary_automatic: Tuple[WeaponRate, ...]
    secondary_automatic: Tuple[WeaponRate, ...]
    primary_semiautomatic: Tuple[WeaponRate, ...]
    secondary_shotguns: Tuple[WeaponRate, ...]
    secondary_gadgets: Tuple[str, ...]

    @property
    def weapons(self) -> Tuple[WeaponRate, ...]:
        """Return all automatic weapons for compatibility and unique counts."""
        return self.primary_automatic + self.secondary_automatic


@dataclass(frozen=True)
class OperatorRating:
    tier: str
    score: float


def normalize_weapon_name(name: str) -> str:
    return WEAPON_NAME_ALIASES.get(name, name)


def operator_key(name: str) -> str:
    """Return the stable ASCII key used by ratings and icon filenames."""
    ascii_name = _ascii_operator_name(name)
    words = []
    current = []
    for character in ascii_name.lower():
        if character.isascii() and character.isalnum():
            current.append(character)
        elif current:
            words.append("".join(current))
            current = []
    if current:
        words.append("".join(current))
    return "-".join(words)


def _ascii_operator_name(name: str) -> str:
    replacements = str.maketrans({"ø": "o", "Ø": "O", "æ": "ae", "Æ": "AE", "ß": "ss"})
    normalized = unicodedata.normalize("NFKD", name.translate(replacements))
    return "".join(character for character in normalized if not unicodedata.combining(character))


def load_ratings(path: Path, operator_names: Iterable[str]) -> Dict[str, OperatorRating]:
    try:
        document = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise DataFormatError("cannot read rating file %s: %s" % (path, error)) from error
    if not isinstance(document, dict):
        raise DataFormatError("rating file must contain a JSON object")
    score_map = document.get("score_map")
    tiers = document.get("tiers")
    if not isinstance(score_map, dict) or not isinstance(tiers, dict):
        raise DataFormatError("rating file must contain score_map and tiers objects")

    ratings: Dict[str, OperatorRating] = {}
    for tier, names in tiers.items():
        score = score_map.get(tier)
        if not isinstance(tier, str) or isinstance(score, bool) or not isinstance(score, (int, float)):
            raise DataFormatError("tier %s has no numeric score" % tier)
        if not isinstance(names, list) or not all(isinstance(name, str) and name for name in names):
            raise DataFormatError("tier %s must contain operator keys" % tier)
        for key in names:
            if key in ratings:
                raise DataFormatError("duplicate rating for operator: %s" % key)
            ratings[key] = OperatorRating(tier=tier, score=score)

    expected = {operator_key(name) for name in operator_names}
    actual = set(ratings)
    if expected != actual:
        missing = sorted(expected - actual)
        extra = sorted(actual - expected)
        raise DataFormatError(
            "rating coverage mismatch; missing=%s; extra=%s"
            % (", ".join(missing) or "none", ", ".join(extra) or "none")
        )
    return ratings


def _json_object(label: str, text: str) -> Mapping[str, object]:
    try:
        value = json.loads(text)
    except (TypeError, json.JSONDecodeError) as error:
        raise DataFormatError("%s is not valid JSON" % label) from error
    if not isinstance(value, dict):
        raise DataFormatError("%s must be a JSON object" % label)
    return value


def _field_names(title: str, fields: object) -> List[str]:
    if not isinstance(fields, list):
        raise DataFormatError("%s schema.fields must be a list" % title)

    names = []
    for field in fields:
        if isinstance(field, str):
            name = field
        elif isinstance(field, dict) and isinstance(field.get("name"), str):
            name = field["name"]
        else:
            raise DataFormatError("%s schema field names must be strings" % title)
        names.append(name)
    if len(set(names)) != len(names):
        raise DataFormatError("%s schema has duplicate field names" % title)
    return names


def parse_tabx_document(
    title: str, document: Mapping[str, object], required_fields: Sequence[str]
) -> List[Dict[str, object]]:
    schema = document.get("schema")
    data = document.get("data")
    if not isinstance(schema, dict):
        raise DataFormatError("%s is missing schema" % title)
    names = _field_names(title, schema.get("fields"))
    missing = [name for name in required_fields if name not in names]
    if missing:
        raise DataFormatError("%s missing required fields: %s" % (title, ", ".join(missing)))
    if not isinstance(data, list):
        raise DataFormatError("%s data must be a list" % title)

    rows = []
    for number, values in enumerate(data, start=1):
        if not isinstance(values, list):
            raise DataFormatError("%s row %d must be a list" % (title, number))
        if len(values) != len(names):
            raise DataFormatError(
                "%s row %d has %d values; expected %d"
                % (title, number, len(values), len(names))
            )
        rows.append(dict(zip(names, values)))
    return rows


def parse_tabx_content(title: str, content: str, required_fields: Sequence[str]) -> List[Dict[str, object]]:
    return parse_tabx_document(title, _json_object(title, content), required_fields)


def _extract_page_content(title: str, document: Mapping[str, object]) -> str:
    query = document.get("query")
    if not isinstance(query, dict):
        raise DataFormatError("%s response is missing query" % title)
    pages = query.get("pages")
    if not isinstance(pages, list) or len(pages) != 1 or not isinstance(pages[0], dict):
        raise DataFormatError("%s response has an invalid pages list" % title)
    revisions = pages[0].get("revisions")
    if not isinstance(revisions, list) or not revisions or not isinstance(revisions[0], dict):
        raise DataFormatError("%s response has no revision content" % title)
    slots = revisions[0].get("slots")
    if not isinstance(slots, dict) or not isinstance(slots.get("main"), dict):
        raise DataFormatError("%s response has no main revision slot" % title)
    content = slots["main"].get("content")
    if not isinstance(content, str):
        raise DataFormatError("%s response main slot has no content" % title)
    return content


def _curl_command(curl_path: str, title: str) -> List[str]:
    return [
        curl_path,
        "--location",
        "--silent",
        "--show-error",
        "--fail",
        "--max-time",
        "30",
        "--user-agent",
        "r6-operator-stats/1.0",
        "--get",
        API_URL,
        "--data-urlencode",
        "action=query",
        "--data-urlencode",
        "prop=revisions",
        "--data-urlencode",
        "rvprop=content",
        "--data-urlencode",
        "rvslots=main",
        "--data-urlencode",
        "format=json",
        "--data-urlencode",
        "formatversion=2",
        "--data-urlencode",
        "titles=%s" % title,
    ]


def fetch_tabx_page(
    title: str,
    required_fields: Sequence[str],
    run_command: Callable[..., object] = subprocess.run,
    sleep: Callable[[float], None] = time.sleep,
    which: Callable[[str], Optional[str]] = shutil.which,
) -> List[Dict[str, object]]:
    curl_path = which("curl.exe") or which("curl")
    if not curl_path:
        raise FetchError("curl.exe or curl was not found on PATH")

    last_error = "unknown curl failure"
    for attempt in range(1, 5):
        try:
            result = run_command(
                _curl_command(curl_path, title),
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="strict",
                check=True,
            )
            output = getattr(result, "stdout", "")
            if not isinstance(output, str) or not output.strip():
                raise ValueError("curl returned empty output")
            outer = _json_object("%s MediaWiki response" % title, output)
            content = _extract_page_content(title, outer)
            tabx = _json_object("%s Tabx content" % title, content)
            return parse_tabx_document(title, tabx, required_fields)
        except subprocess.CalledProcessError as error:
            last_error = "curl failed: %s" % (error.stderr or error)
        except OSError as error:
            last_error = "curl failed: %s" % error
        except ValueError as error:
            last_error = str(error)
        except DataFormatError as error:
            if "not valid JSON" not in str(error):
                raise FetchError(str(error)) from error
            last_error = str(error)

        if attempt < 4:
            sleep(float(attempt))
    raise FetchError("failed to fetch %s after 4 attempts: %s" % (title, last_error))


def _required_string(record: Mapping[str, object], field: str, label: str) -> str:
    value = record.get(field)
    if not isinstance(value, str) or not value:
        raise DataFormatError("%s %s must be a nonempty string" % (label, field))
    return value


def _numeric(record: Mapping[str, object], field: str, label: str) -> float:
    value = record.get(field)
    if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(value):
        raise DataFormatError("%s %s must be numeric" % (label, field))
    return value


def _plain_wiki_text(value: str) -> str:
    value = re.sub(
        r"\[\[([^\[\]]+)\]\]",
        lambda match: match.group(1).split("|")[-1],
        value,
    )
    return value.replace("[", "").replace("]", "").strip()


def build_operator_rows(
    operators: Iterable[Mapping[str, object]],
    weapons: Iterable[Mapping[str, object]],
    configs: Iterable[Mapping[str, object]],
) -> Dict[str, List[OperatorRow]]:
    operator_data = {}
    for record in operators:
        operator_id = _required_string(record, "id", "operator")
        if operator_id in operator_data:
            raise DataFormatError("duplicate operator id: %s" % operator_id)
        camp = _required_string(record, "camp", "operator")
        if camp not in SIDES:
            raise DataFormatError("unknown camp: %s" % camp)
        props = _required_string(record, "props", "operator")
        gadgets = tuple(
            cleaned
            for part in props.replace("；", ";").split(";")
            if (cleaned := _plain_wiki_text(part))
        )
        if not gadgets:
            raise DataFormatError("operator props must contain at least one secondary gadget")
        operator_data[operator_id] = (
            _required_string(record, "name", "operator"),
            camp,
            _numeric(record, "speed", "operator"),
            _numeric(record, "_index", "operator"),
            gadgets,
        )

    weapon_data = {}
    for record in weapons:
        weapon_id = _required_string(record, "id", "weapon")
        if weapon_id in weapon_data:
            raise DataFormatError("duplicate weapon id: %s" % weapon_id)
        equipment = _numeric(record, "equipment", "weapon")
        if equipment not in (1, 2):
            raise DataFormatError("weapon equipment must be 1 or 2: %s" % weapon_id)
        weapon_data[weapon_id] = (
            normalize_weapon_name(
                _required_string(record, "zh_model", "weapon")
            ),
            _numeric(record, "firerate", "weapon"),
            _numeric(record, "projectile", "weapon"),
            _numeric(record, "index", "weapon"),
            _required_string(record, "type", "weapon"),
            int(equipment),
        )

    grouped_by_operator = {
        operator_id: {
            "primary_automatic": [],
            "secondary_automatic": [],
            "primary_semiautomatic": [],
            "secondary_shotguns": [],
        }
        for operator_id in operator_data
    }
    seen_pairs = set()
    for record in configs:
        weapon_id = _required_string(record, "id", "weapon config")
        operator_id = _required_string(record, "user", "weapon config")
        if operator_id not in operator_data:
            raise DataFormatError("weapon config references unknown operator: %s" % operator_id)
        if weapon_id not in weapon_data:
            raise DataFormatError("weapon config references unknown weapon: %s" % weapon_id)
        pair = (operator_id, weapon_id)
        if pair in seen_pairs:
            continue
        seen_pairs.add(pair)

        name, firerate, projectile, weapon_order, _weapon_type, equipment = weapon_data[weapon_id]
        rate = WeaponRate(name=name, firerate=firerate, order=weapon_order, weapon_id=weapon_id)
        is_automatic = firerate > 0 and projectile == 1
        if equipment == 1:
            if is_automatic:
                grouped_by_operator[operator_id]["primary_automatic"].append(rate)
            elif firerate == 0 and projectile == 1:
                grouped_by_operator[operator_id]["primary_semiautomatic"].append(rate)
        elif is_automatic:
            grouped_by_operator[operator_id]["secondary_automatic"].append(rate)
        elif projectile > 1:
            grouped_by_operator[operator_id]["secondary_shotguns"].append(rate)

    rows = {camp: [] for camp in SIDES}
    for operator_id, (name, camp, speed, operator_order, gadgets) in operator_data.items():
        groups = grouped_by_operator[operator_id]
        primary_automatic = tuple(
            sorted(groups["primary_automatic"], key=lambda rate: (-rate.firerate, rate.order))
        )
        secondary_automatic = tuple(
            sorted(groups["secondary_automatic"], key=lambda rate: (-rate.firerate, rate.order))
        )
        primary_semiautomatic = tuple(
            sorted(groups["primary_semiautomatic"], key=lambda rate: rate.order)
        )
        secondary_shotguns = tuple(
            sorted(groups["secondary_shotguns"], key=lambda rate: rate.order)
        )
        rows[camp].append(
            OperatorRow(
                name=name,
                speed=speed,
                order=operator_order,
                primary_automatic=primary_automatic,
                secondary_automatic=secondary_automatic,
                primary_semiautomatic=primary_semiautomatic,
                secondary_shotguns=secondary_shotguns,
                secondary_gadgets=gadgets,
            )
        )
    for camp in SIDES:
        rows[camp].sort(key=lambda row: (-row.speed, row.order))
    return rows


def fetch_all() -> Dict[str, List[OperatorRow]]:
    operators = fetch_tabx_page("Data:Operator.tabx", OPERATOR_FIELDS)
    weapons = fetch_tabx_page("Data:WeaponData.tabx", WEAPON_FIELDS)
    configs = fetch_tabx_page("Data:WeaponConfig.tabx", CONFIG_FIELDS)
    return build_operator_rows(operators, weapons, configs)


def load_snapshot_rows(data_dir: Path) -> Dict[str, List[OperatorRow]]:
    """Load the validated three-table Wiki snapshot without network access."""
    root = Path(data_dir) / "wiki"
    tables = {}
    for key, _title, required_fields in (
        ("operator", "Data:Operator.tabx", OPERATOR_FIELDS),
        ("weapon", "Data:WeaponData.tabx", WEAPON_FIELDS),
        ("weapon_config", "Data:WeaponConfig.tabx", CONFIG_FIELDS),
    ):
        path = root / ("%s.json" % key)
        try:
            document = json.loads(path.read_text(encoding="utf-8"))
        except OSError as error:
            raise DataFormatError(
                "cannot read Wiki snapshot %s: %s" % (path, error)
            ) from error
        except json.JSONDecodeError as error:
            raise DataFormatError(
                "invalid JSON in Wiki snapshot %s: %s" % (path, error)
            ) from error
        if not isinstance(document, dict) or document.get("schema_version") != 1:
            raise DataFormatError(
                "Wiki snapshot schema_version must be 1: %s" % path
            )
        fields = document.get("fields")
        if fields != list(required_fields):
            raise DataFormatError(
                "Wiki snapshot fields changed: %s" % path
            )
        records = document.get("records")
        if not isinstance(records, list) or not all(
            isinstance(record, dict) for record in records
        ):
            raise DataFormatError(
                "Wiki snapshot records must be objects: %s" % path
            )
        tables[key] = records
    return build_operator_rows(
        tables["operator"],
        tables["weapon"],
        tables["weapon_config"],
    )


def count_unique_automatic_weapons(rows: Mapping[str, Iterable[OperatorRow]]) -> int:
    return len(
        {
            rate.weapon_id or (row.name, rate.name, rate.firerate, rate.order)
            for camp in SIDES
            for row in rows[camp]
            for rate in row.weapons
        }
    )


def _run_curl_json(parameters: Mapping[str, str]) -> Mapping[str, object]:
    curl_path = shutil.which("curl.exe") or shutil.which("curl")
    if not curl_path:
        raise FetchError("curl.exe or curl was not found on PATH")
    url = "https://r6s.huijiwiki.com/api.php?" + urlencode(parameters)
    last_error = "unknown curl failure"
    for attempt in range(1, 5):
        try:
            result = subprocess.run(
                [
                    curl_path,
                    "--location",
                    "--silent",
                    "--show-error",
                    "--fail",
                    "--max-time",
                    "30",
                    "--user-agent",
                    "r6-operator-stats/1.0",
                    url,
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="strict",
                check=True,
            )
            if not result.stdout.strip():
                raise ValueError("curl returned empty output")
            return _json_object("Huiji Wiki API response", result.stdout)
        except (subprocess.CalledProcessError, OSError, ValueError, DataFormatError) as error:
            last_error = str(error)
            if attempt < 4:
                time.sleep(float(attempt))
    raise FetchError("Huiji Wiki API failed after 4 attempts: %s" % last_error)


def _wiki_icon_url(operator_name: str) -> str:
    key = operator_key(operator_name)
    skill_candidates = []
    badge_candidates = []
    for page_title in (_ascii_operator_name(operator_name), "特勤干员"):
        page_result = _run_curl_json(
            {
                "action": "query",
                "titles": page_title,
                "prop": "images",
                "imlimit": "max",
                "format": "json",
                "formatversion": "2",
            }
        )
        pages = page_result.get("query", {}).get("pages", []) if isinstance(page_result.get("query"), dict) else []
        if not isinstance(pages, list) or len(pages) != 1 or not isinstance(pages[0], dict):
            raise DataFormatError("invalid image list for operator %s" % operator_name)
        images = pages[0].get("images", [])
        if not isinstance(images, list):
            raise DataFormatError("invalid images field for operator %s" % operator_name)
        for image in images:
            title = image.get("title") if isinstance(image, dict) else None
            if not isinstance(title, str) or ":" not in title:
                continue
            filename = title.split(":", 1)[1]
            lower = filename.lower()
            prefix = filename[: lower.rfind(" skill icon")]
            same_name = operator_key(prefix).replace("-", "") == key.replace("-", "")
            if " skill icon" in lower and " old" not in lower and same_name:
                skill_candidates.append(title)
            if lower.startswith("badge ") and lower.endswith(".png"):
                badge_name = filename[6:-4]
                if operator_key(badge_name).replace("-", "") == key.replace("-", ""):
                    badge_candidates.append(title)
        if skill_candidates or badge_candidates:
            break
    candidates = list(dict.fromkeys(skill_candidates + badge_candidates))
    if not candidates:
        raise DataFormatError(
            "no Huiji Wiki skill or badge icon found for %s" % operator_name
        )
    for candidate in candidates:
        info_result = _run_curl_json(
            {
                "action": "query",
                "titles": candidate,
                "prop": "imageinfo",
                "iiprop": "url",
                "format": "json",
                "formatversion": "2",
            }
        )
        info_pages = info_result.get("query", {}).get("pages", []) if isinstance(info_result.get("query"), dict) else []
        try:
            url = info_pages[0]["imageinfo"][0]["url"]
        except (IndexError, KeyError, TypeError):
            continue
        if isinstance(url, str) and url.startswith("https://huiji-public.huijistatic.com/r6s/"):
            return url
    raise DataFormatError("Huiji Wiki icon URL is missing for %s" % operator_name)


def _wiki_badge_url(operator_name: str) -> str:
    key = operator_key(operator_name).replace("-", "")
    candidates = []
    for page_title in (_ascii_operator_name(operator_name), "特勤干员"):
        page_result = _run_curl_json(
            {
                "action": "query",
                "titles": page_title,
                "prop": "images",
                "imlimit": "max",
                "format": "json",
                "formatversion": "2",
            }
        )
        pages = page_result.get("query", {}).get("pages", []) if isinstance(page_result.get("query"), dict) else []
        if not isinstance(pages, list) or len(pages) != 1 or not isinstance(pages[0], dict):
            raise DataFormatError("invalid badge list for operator %s" % operator_name)
        images = pages[0].get("images", [])
        if not isinstance(images, list):
            raise DataFormatError("invalid images field for operator %s" % operator_name)
        for image in images:
            title = image.get("title") if isinstance(image, dict) else None
            if not isinstance(title, str) or ":" not in title:
                continue
            filename = title.split(":", 1)[1]
            lower = filename.lower()
            if lower.startswith("badge ") and lower.endswith(".png"):
                badge_name = filename[6:-4]
                if operator_key(badge_name).replace("-", "") == key:
                    candidates.append(title)
        if candidates:
            break
    candidates = list(dict.fromkeys(candidates))
    if not candidates:
        raise DataFormatError("no Huiji Wiki badge found for %s" % operator_name)

    for candidate in candidates:
        info_result = _run_curl_json(
            {
                "action": "query",
                "titles": candidate,
                "prop": "imageinfo",
                "iiprop": "url",
                "format": "json",
                "formatversion": "2",
            }
        )
        info_pages = info_result.get("query", {}).get("pages", []) if isinstance(info_result.get("query"), dict) else []
        try:
            url = info_pages[0]["imageinfo"][0]["url"]
        except (IndexError, KeyError, TypeError):
            continue
        if isinstance(url, str) and url.startswith("https://huiji-public.huijistatic.com/r6s/"):
            return url
    raise DataFormatError("Huiji Wiki badge URL is missing for %s" % operator_name)


def prepare_operator_icons(
    rows: Mapping[str, Sequence[OperatorRow]], icon_dir: Path
) -> Path:
    icon_dir.mkdir(parents=True, exist_ok=True)
    white_dir = icon_dir / "white"
    badge_dir = icon_dir / "badge"
    legacy_raw_dir = icon_dir / "_raw"
    legacy_badge_dir = icon_dir / "_badges"
    white_dir.mkdir(parents=True, exist_ok=True)
    badge_dir.mkdir(parents=True, exist_ok=True)
    curl_path = shutil.which("curl.exe") or shutil.which("curl")
    if not curl_path:
        raise FetchError("curl.exe or curl was not found on PATH")
    all_rows = [row for camp in SIDES for row in rows[camp]]
    for number, row in enumerate(all_rows, start=1):
        filename = operator_key(row.name) + ".png"
        white_destination = white_dir / filename
        badge_destination = badge_dir / filename
        legacy_badge = legacy_badge_dir / filename

        if not badge_destination.exists() and legacy_badge.exists():
            shutil.copyfile(legacy_badge, badge_destination)
        if not badge_destination.exists():
            print("下载灰机 Wiki Badge %d/%d：%s" % (number, len(all_rows), operator_key(row.name)))
            badge_url = _wiki_badge_url(row.name)
            try:
                subprocess.run(
                    [
                        curl_path,
                        "--location",
                        "--silent",
                        "--show-error",
                        "--fail",
                        "--max-time",
                        "30",
                        "--user-agent",
                        "r6-operator-stats/1.0",
                        "--output",
                        str(badge_destination),
                        badge_url,
                    ],
                    check=True,
                )
                with PillowImage.open(badge_destination) as image:
                    image.verify()
            except (subprocess.CalledProcessError, OSError) as error:
                badge_destination.unlink(missing_ok=True)
                raise FetchError("failed to download badge for %s: %s" % (row.name, error)) from error

        if not white_destination.exists():
            legacy_raw = legacy_raw_dir / filename
            if legacy_raw.exists():
                make_white_operator_icon(legacy_raw, white_destination)
            else:
                print("下载灰机 Wiki 白色图案 %d/%d：%s" % (number, len(all_rows), operator_key(row.name)))
                with tempfile.TemporaryDirectory() as directory:
                    source = Path(directory) / filename
                    url = _wiki_icon_url(row.name)
                    try:
                        subprocess.run(
                            [
                                curl_path,
                                "--location",
                                "--silent",
                                "--show-error",
                                "--fail",
                                "--max-time",
                                "30",
                                "--user-agent",
                                "r6-operator-stats/1.0",
                                "--output",
                                str(source),
                                url,
                            ],
                            check=True,
                        )
                        make_white_operator_icon(source, white_destination)
                    except (subprocess.CalledProcessError, OSError) as error:
                        white_destination.unlink(missing_ok=True)
                        raise FetchError("failed to download white icon for %s: %s" % (row.name, error)) from error
    return icon_dir


def make_white_operator_icon(source: Path, destination: Path) -> None:
    try:
        with PillowImage.open(source) as image:
            rgba = image.convert("RGBA")
    except OSError as error:
        raise DataFormatError("invalid operator icon: %s" % source) from error
    red, green, blue, alpha = rgba.split()
    brightest = ImageChops.lighter(ImageChops.lighter(red, green), blue)
    visible_color = brightest.point(lambda value: 255 if value >= 32 else 0)
    alpha = ImageChops.multiply(alpha, visible_color)
    bounds = alpha.getbbox()
    if bounds is None:
        raise DataFormatError("operator icon is fully transparent: %s" % source)
    alpha = alpha.crop(bounds)
    alpha.thumbnail((72, 72), PillowImage.Resampling.LANCZOS)

    canvas = PillowImage.new("RGBA", (96, 96), (0, 0, 0, 0))
    pattern = PillowImage.new("RGBA", alpha.size, (255, 255, 255, 255))
    pattern.putalpha(alpha)
    position = ((96 - alpha.width) // 2, (96 - alpha.height) // 2)
    canvas.alpha_composite(pattern, position)
    canvas.save(destination, format="PNG", optimize=True)


def _format_rate(value: float) -> str:
    return str(int(value)) if float(value).is_integer() else str(value)


def _format_automatic_weapons(rates: Sequence[WeaponRate]) -> str:
    return "\n".join(
        "%s（%s）" % (rate.name, _format_rate(rate.firerate)) for rate in rates
    ) or "无自动枪械"


def _format_weapon_names(rates: Sequence[WeaponRate]) -> str:
    return "\n".join(rate.name for rate in rates) or "无"


def write_workbook(
    path: Path,
    rows: Mapping[str, Sequence[OperatorRow]],
    ratings: Mapping[str, OperatorRating],
    icon_dir: Path,
    report_sources: ReportSources,
) -> None:
    workbook = Workbook()
    first_sheet = workbook.active
    for index, camp in enumerate(SIDES):
        sheet = first_sheet if index == 0 else workbook.create_sheet()
        sheet.title = camp
        sheet.append(HEADERS)
        for row in rows[camp]:
            key = operator_key(row.name)
            rating = ratings.get(key)
            if rating is None:
                raise DataFormatError("missing rating for operator: %s" % row.name)
            icon_path = icon_dir / "badge" / (key + ".png")
            if not icon_path.is_file():
                raise DataFormatError("missing icon for operator: %s" % row.name)
            values = (
                _format_automatic_weapons(row.primary_automatic),
                _format_automatic_weapons(row.secondary_automatic),
                _format_weapon_names(row.primary_semiautomatic),
                _format_weapon_names(row.secondary_shotguns),
                "\n".join(row.secondary_gadgets),
            )
            sheet.append((None, row.name, row.speed) + values + (rating.score,))
            row_number = sheet.max_row
            icon = ExcelImage(str(icon_path))
            icon.width = 34
            icon.height = 34
            sheet.add_image(icon, "A%d" % row_number)
            max_lines = max(value.count("\n") + 1 for value in values)
            sheet.row_dimensions[row_number].height = max(31.5, max_lines * 15)

        sheet.freeze_panes = "A2"
        sheet.row_dimensions[1].height = 24
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        rating_source = report_sources.rating
        rating_note = (
            "来源：%s《%s》\n%s\n发布日期：%s；最终榜单画面：%s\n"
            "原始分数：S=100，A=85，B=70，C=55，D=40，F=20，"
            "boof=0；boof 展示为最低字母档 F。"
            % (
                rating_source.creator,
                rating_source.title,
                rating_source.url,
                rating_source.published.isoformat(),
                rating_source.final_frame,
            )
        )
        sheet["I1"].comment = Comment(rating_note, "Codex")
        for row_number in range(2, sheet.max_row + 1):
            sheet.cell(row_number, 1).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row_number, 2).alignment = Alignment(vertical="center")
            sheet.cell(row_number, 3).alignment = Alignment(horizontal="center", vertical="center")
            for column in range(4, 9):
                sheet.cell(row_number, column).alignment = Alignment(wrap_text=True, vertical="center")
            sheet.cell(row_number, 9).alignment = Alignment(horizontal="center", vertical="center")
            raw_tier = ratings[operator_key(str(sheet.cell(row_number, 2).value))].tier
            try:
                tier = display_tier(raw_tier)
            except ValueError as exc:
                raise DataFormatError(str(exc)) from exc
            fill_color = TIER_COLORS[tier]
            sheet.cell(row_number, 9).fill = PatternFill("solid", fgColor=fill_color)
            sheet.cell(row_number, 9).font = Font(
                bold=True, color="FFFFFF" if tier in ("S", "F") else "000000"
            )
        for column, width in zip("ABCDEFGHI", (8, 18, 10, 34, 32, 24, 20, 32, 14)):
            sheet.column_dimensions[column].width = width
        last_data_row = sheet.max_row
        table_ref = "A1:I%d" % last_data_row
        table = Table(displayName="%sOperators" % ("Attack" if camp == ATTACKERS else "Defense"), ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
        )
        sheet.add_table(table)
        _, footer_end = append_source_footer(
            sheet, len(HEADERS), report_sources
        )
        sheet.print_area = "A1:I%d" % footer_end

    scores = {}
    for patch in report_sources.patches:
        for change in patch.changes:
            rating = ratings.get(operator_key(change.subject))
            if rating is None:
                raise DataFormatError(
                    "missing rating for patch subject: %s" % change.subject
                )
            scores[change.subject] = int(rating.score)
    add_patch_notes_sheet(workbook, scores, report_sources)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main(
    argv: Optional[Sequence[str]] = None,
    fetcher: Optional[Callable[[], Dict[str, List[OperatorRow]]]] = None,
    rating_loader: Callable[[Path, Iterable[str]], Dict[str, OperatorRating]] = load_ratings,
    icon_preparer: Callable[[Mapping[str, Sequence[OperatorRow]], Path], Path] = prepare_operator_icons,
    source_loader: Callable[[Path], ReportSources] = load_report_sources,
) -> int:
    parser = argparse.ArgumentParser(description="Export Rainbow Six operator automatic weapon statistics.")
    parser.add_argument("--inputs-dir", type=Path, default=DEFAULT_INPUTS_DIR)
    parser.add_argument(
        "--output",
        default=str(Path("~temp") / "r6_operator_stats.xlsx"),
        help="output XLSX path",
    )
    parser.add_argument("--ratings", type=Path, default=None)
    parser.add_argument("--icons-dir", type=Path, default=None)
    args = parser.parse_args(argv)
    output = Path(args.output)
    try:
        rows = fetcher() if fetcher is not None else load_snapshot_rows(args.inputs_dir)
        operator_names = [row.name for camp in SIDES for row in rows[camp]]
        ratings_path = args.ratings or args.inputs_dir / "athieno" / "latest.json"
        icons_path = args.icons_dir or args.inputs_dir / "icons" / "operator"
        ratings = rating_loader(ratings_path, operator_names)
        icon_dir = (
            icon_preparer(rows, icons_path)
            if fetcher is not None
            else icons_path
        )
        report_sources = source_loader(args.inputs_dir)
        write_workbook(output, rows, ratings, icon_dir, report_sources)
    except (OSError, R6StatsError, SourceDataError, ValueError) as error:
        print("错误：%s" % error, file=sys.stderr)
        return 1

    resolved = output.resolve()
    print("进攻方干员：%d" % len(rows[ATTACKERS]))
    print("防守方干员：%d" % len(rows[DEFENDERS]))
    print("唯一自动枪械：%d" % count_unique_automatic_weapons(rows))
    print("输出文件：%s" % resolved)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
