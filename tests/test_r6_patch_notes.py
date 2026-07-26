import unittest

from openpyxl import Workbook

import _path_setup
from r6_report import patch_notes as notes
from r6_report.workbook_sources import append_source_footer
from source_fixtures import make_report_sources


class PatchNotesTests(unittest.TestCase):
    def score_map(self):
        return {"Alice": 70, "Bob": 70}

    def test_appends_three_source_rows_as_last_nonempty_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "content"

        start, end = append_source_footer(sheet, 9, make_report_sources())

        self.assertEqual((start, end), (3, 5))
        self.assertEqual(sheet.max_row, 5)
        self.assertIn("评分来源：Athieno", sheet["A3"].value)
        self.assertIn("游戏数据：灰机 Wiki", sheet["A4"].value)
        self.assertIn("补丁区间：", sheet["A5"].value)
        self.assertIn("A3:I3", {str(item) for item in sheet.merged_cells.ranges})
        self.assertTrue(sheet["A3"].alignment.wrap_text)
        self.assertEqual(sheet["A3"].hyperlink.target, "https://youtu.be/fAjTjhNdJe4")

    def test_builds_patch_sheet_with_sources_directions_and_scores(self):
        workbook = Workbook()
        sheet = notes.add_patch_notes_sheet(
            workbook, self.score_map(), make_report_sources(with_changes=True)
        )

        self.assertEqual(sheet.title, "补丁说明")
        self.assertIsNone(sheet.freeze_panes)
        self.assertIn("除 Athieno Y11S2 视频评分外", sheet["A2"].value)
        values = [cell.value for row in sheet.iter_rows() for cell in row]
        self.assertIn("Y11S2.1", values)
        self.assertIn("Y11S2.2", values)
        self.assertIn("Alice", values)
        self.assertIn("Bob", values)
        self.assertIn("B / 70", values)
        self.assertGreaterEqual(sheet.column_dimensions["A"].width, 16)

        direction_fills = {}
        for row in range(1, sheet.max_row + 1):
            direction = sheet.cell(row, 1).value
            if direction in notes.DIRECTION_COLOURS:
                direction_fills.setdefault(
                    direction, sheet.cell(row, 1).fill.fgColor.rgb[-6:]
                )
        self.assertEqual(direction_fills, notes.DIRECTION_COLOURS)
        self.assertEqual(
            notes.DIRECTION_COLOURS,
            {
                "增强": "548235",
                "削弱": "C00000",
                "混合": "BF9000",
            },
        )
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row, 1).value in notes.DIRECTION_COLOURS:
                self.assertTrue(sheet.cell(row, 1).font.bold)
                self.assertEqual(
                    sheet.cell(row, 1).font.color.rgb[-6:],
                    "FFFFFF",
                )
        self.assertIsNone(sheet.freeze_panes)
        self.assertIn("2026-07-25", sheet.cell(sheet.max_row, 1).value)

    def test_rejects_missing_video_score(self):
        with self.assertRaisesRegex(notes.PatchNotesError, "missing video score"):
            notes.add_patch_notes_sheet(
                Workbook(), {}, make_report_sources(with_changes=True)
            )

    def test_rejects_existing_patch_sheet(self):
        workbook = Workbook()
        workbook.create_sheet("补丁说明")

        with self.assertRaisesRegex(notes.PatchNotesError, "already exists"):
            notes.add_patch_notes_sheet(
                workbook, self.score_map(), make_report_sources(with_changes=True)
            )


if __name__ == "__main__":
    unittest.main()
