import contextlib
import io
import json
import subprocess
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
from PIL import Image

import _path_setup
from r6_report import operator_stats as stats
from r6_report import patch_notes as notes
from source_fixtures import make_report_sources


OPERATOR_FIELDS = ("id", "name", "camp", "speed", "_index", "props")
WEAPON_FIELDS = ("id", "zh_model", "firerate", "projectile", "index", "type", "equipment")
CONFIG_FIELDS = ("id", "user")


def tabx(fields, rows):
    return json.dumps(
        {
            "schema": {"fields": [{"name": field, "type": "string"} for field in fields]},
            "data": rows,
        }
    )


def operator(operator_id, name, camp, speed, index, props="手雷×2;烟雾弹×2"):
    return {
        "id": operator_id,
        "name": name,
        "camp": camp,
        "speed": speed,
        "_index": index,
        "props": props,
    }


def weapon(weapon_id, name, firerate, projectile, index, weapon_type="突击步枪", equipment=1):
    return {
        "id": weapon_id,
        "zh_model": name,
        "firerate": firerate,
        "projectile": projectile,
        "index": index,
        "type": weapon_type,
        "equipment": equipment,
    }


def mediawiki_response(content):
    return json.dumps(
        {
            "query": {
                "pages": [
                    {
                        "title": "Data:Example.tabx",
                        "revisions": [{"slots": {"main": {"content": content}}}],
                    }
                ]
            }
        }
    )


class ParseTabxTests(unittest.TestCase):
    def test_parses_schema_fields_and_rejects_missing_fields_or_row_widths(self):
        content = tabx(OPERATOR_FIELDS, [["op1", "Alice", "进攻方", 3, 2, "手雷×2;烟雾弹×2"]])
        rows = stats.parse_tabx_content("Data:Operator.tabx", content, OPERATOR_FIELDS)
        self.assertEqual(rows, [operator("op1", "Alice", "进攻方", 3, 2)])

        with self.assertRaisesRegex(stats.DataFormatError, "missing required fields"):
            stats.parse_tabx_content(
                "Data:Operator.tabx",
                tabx(("id", "name"), [["op1", "Alice"]]),
                OPERATOR_FIELDS,
            )
        with self.assertRaisesRegex(stats.DataFormatError, "row 1 has 2 values"):
            stats.parse_tabx_content(
                "Data:Operator.tabx",
                tabx(OPERATOR_FIELDS, [["op1", "Alice"]]),
                OPERATOR_FIELDS,
            )


class JoinTests(unittest.TestCase):
    def test_auto_filter_is_exactly_positive_numeric_firerate_and_projectile_one(self):
        operators = [operator("op", "Alice", "进攻方", 3, 1)]
        weapons = [
            weapon("positive", "Positive", 1, 1, 1),
            weapon("zero-rate", "Zero rate", 0, 1, 2),
            weapon("negative-rate", "Negative rate", -1, 1, 3),
            weapon("other-projectile", "Other projectile", 900, 2, 4),
            weapon("manual", "Manual", 900, 0, 5),
        ]
        configs = [{"id": item["id"], "user": "op"} for item in weapons]

        rows = stats.build_operator_rows(operators, weapons, configs)
        self.assertEqual([rate.name for rate in rows["进攻方"][0].weapons], ["Positive"])

    def test_joins_deduplicates_splits_and_sorts_rows_and_weapons(self):
        operators = [
            operator("slow", "Slow", "进攻方", 2, 9),
            operator("fast-later", "Fast later", "进攻方", 3, 8),
            operator("fast-first", "Fast first", "进攻方", 3, 1),
            operator("def", "Defender", "防守方", 1, 1),
        ]
        weapons = [
            weapon("low", "Low", 700, 1, 1),
            weapon("high-later", "High later", 900, 1, 4),
            weapon("high-first", "High first", 900, 1, 2),
            weapon("def-gun", "Def gun", 800, 1, 1),
        ]
        configs = [
            {"id": "low", "user": "slow"},
            {"id": "high-later", "user": "slow"},
            {"id": "high-first", "user": "slow"},
            {"id": "high-first", "user": "slow"},
            {"id": "def-gun", "user": "def"},
        ]

        rows = stats.build_operator_rows(operators, weapons, configs)
        self.assertEqual([row.name for row in rows["进攻方"]], ["Fast first", "Fast later", "Slow"])
        self.assertEqual([row.name for row in rows["防守方"]], ["Defender"])
        self.assertEqual(
            [rate.name for rate in rows["进攻方"][2].weapons],
            ["High first", "High later", "Low"],
        )
        self.assertEqual(stats.count_unique_automatic_weapons(rows), 4)

    def test_includes_operators_without_automatic_weapons(self):
        rows = stats.build_operator_rows(
            [operator("op", "No auto", "防守方", 2, 1)],
            [weapon("manual", "Manual", 600, 0, 1)],
            [{"id": "manual", "user": "op"}],
        )
        self.assertEqual(rows["防守方"][0].primary_automatic, ())
        self.assertEqual(rows["防守方"][0].secondary_automatic, ())

    def test_normalizes_roni_conversion_variant(self):
        self.assertEqual(
            stats.normalize_weapon_name("P10 RONI转换套件衍生型"),
            "P10 RONI",
        )
        self.assertEqual(
            stats.normalize_weapon_name("Commando 9"),
            "Commando 9",
        )

        rows = stats.build_operator_rows(
            [operator("op", "Mozzie", "防守方", 2, 1)],
            [
                weapon(
                    "roni",
                    "P10 RONI转换套件衍生型",
                    980,
                    1,
                    1,
                )
            ],
            [{"id": "roni", "user": "op"}],
        )

        self.assertEqual(
            rows["防守方"][0].primary_automatic[0].name,
            "P10 RONI",
        )

    def test_splits_primary_secondary_semiautomatic_shotguns_and_gadgets(self):
        weapons = [
            weapon("primary-auto", "Primary Auto", 800, 1, 1),
            weapon("secondary-auto", "Secondary Auto", 1000, 1, 2, "自动手枪", 2),
            weapon("dmr", "DMR", 0, 1, 3, "精准射手步枪", 1),
            weapon("slug-shotgun", "Slug Shotgun", 0, 1, 4, "霰弹枪", 1),
            weapon("primary-shotgun", "Primary Shotgun", 0, 8, 5, "霰弹枪", 1),
            weapon("shorty", "Shorty", 0, 8, 6, "霰弹枪", 2),
            weapon("bailiff", "Bailiff 410", 0, 4, 7, "手枪", 2),
            weapon("secondary-slug", "Secondary Slug", 0, 1, 8, "霰弹枪", 2),
            weapon("pistol", "Pistol", 0, 1, 9, "手枪", 2),
        ]
        rows = stats.build_operator_rows(
            [
                operator(
                    "op",
                    "Alice",
                    "进攻方",
                    2,
                    1,
                    "手雷×2;[[装备页|烟雾弹]]×2;电磁脉冲式冲击弹]]×2",
                )
            ],
            weapons,
            [{"id": item["id"], "user": "op"} for item in weapons],
        )
        row = rows["进攻方"][0]
        self.assertEqual([rate.name for rate in row.primary_automatic], ["Primary Auto"])
        self.assertEqual([rate.name for rate in row.secondary_automatic], ["Secondary Auto"])
        self.assertEqual(
            [rate.name for rate in row.primary_semiautomatic], ["DMR", "Slug Shotgun"]
        )
        self.assertEqual(
            [rate.name for rate in row.secondary_shotguns], ["Shorty", "Bailiff 410"]
        )
        self.assertEqual(
            row.secondary_gadgets,
            ("手雷×2", "烟雾弹×2", "电磁脉冲式冲击弹×2"),
        )

    def test_rejects_unknown_camps_dangling_configs_and_invalid_numeric_values(self):
        valid_operator = [operator("op", "Alice", "进攻方", 3, 1)]
        valid_weapon = [weapon("gun", "Gun", 900, 1, 1)]
        with self.assertRaisesRegex(stats.DataFormatError, "unknown camp"):
            stats.build_operator_rows(
                [operator("op", "Alice", "观战方", 3, 1)], valid_weapon, [{"id": "gun", "user": "op"}]
            )
        with self.assertRaisesRegex(stats.DataFormatError, "unknown operator"):
            stats.build_operator_rows(valid_operator, valid_weapon, [{"id": "gun", "user": "missing"}])
        with self.assertRaisesRegex(stats.DataFormatError, "unknown weapon"):
            stats.build_operator_rows(valid_operator, valid_weapon, [{"id": "missing", "user": "op"}])

        invalid_cases = [
            ([operator("op", "Alice", "进攻方", True, 1)], valid_weapon),
            (valid_operator, [weapon("gun", "Gun", "900", 1, 1)]),
            (valid_operator, [weapon("gun", "Gun", 900, True, 1)]),
            (valid_operator, [weapon("gun", "Gun", 900, 1, "1")]),
        ]
        for operators, weapons in invalid_cases:
            with self.subTest(operators=operators, weapons=weapons):
                with self.assertRaisesRegex(stats.DataFormatError, "must be numeric"):
                    stats.build_operator_rows(operators, weapons, [{"id": "gun", "user": "op"}])

        with self.assertRaisesRegex(stats.DataFormatError, "equipment must be 1 or 2"):
            stats.build_operator_rows(
                valid_operator,
                [weapon("gun", "Gun", 900, 1, 1, equipment=3)],
                [{"id": "gun", "user": "op"}],
            )
        with self.assertRaisesRegex(stats.DataFormatError, "props must be a nonempty string"):
            stats.build_operator_rows(
                [operator("op", "Alice", "进攻方", 3, 1, props="")],
                valid_weapon,
                [{"id": "gun", "user": "op"}],
            )


class FetchTests(unittest.TestCase):
    def test_retries_curl_empty_and_invalid_json_then_succeeds(self):
        calls = []
        sleeps = []
        responses = [
            subprocess.CalledProcessError(22, ["curl.exe"], stderr="bad gateway"),
            SimpleNamespace(stdout="", stderr=""),
            SimpleNamespace(stdout="not-json", stderr=""),
            SimpleNamespace(
                stdout=mediawiki_response(tabx(("id",), [["ok"]])), stderr=""
            ),
        ]

        def runner(command, **kwargs):
            calls.append((command, kwargs))
            result = responses.pop(0)
            if isinstance(result, Exception):
                raise result
            return result

        rows = stats.fetch_tabx_page(
            "Data:Example.tabx",
            ("id",),
            run_command=runner,
            sleep=sleeps.append,
            which=lambda name: "C:/curl.exe" if name == "curl.exe" else None,
        )
        self.assertEqual(rows, [{"id": "ok"}])
        self.assertEqual(len(calls), 4)
        self.assertEqual(sleeps, [1, 2, 3])
        command, kwargs = calls[0]
        self.assertEqual(command[0], "C:/curl.exe")
        self.assertIn("--location", command)
        max_time_index = command.index("--max-time")
        self.assertEqual(command[max_time_index + 1], "30")
        self.assertEqual(
            kwargs,
            {
                "capture_output": True,
                "text": True,
                "encoding": "utf-8",
                "errors": "strict",
                "check": True,
            },
        )

    def test_exhausts_four_attempts_with_a_clear_error(self):
        calls = []

        def runner(command, **kwargs):
            calls.append(command)
            return SimpleNamespace(stdout="{broken", stderr="")

        with self.assertRaisesRegex(stats.FetchError, "after 4 attempts"):
            stats.fetch_tabx_page(
                "Data:Example.tabx",
                ("id",),
                run_command=runner,
                sleep=lambda seconds: None,
                which=lambda name: "C:/curl.exe",
            )
        self.assertEqual(len(calls), 4)

    def test_retries_oserror_then_succeeds(self):
        calls = []
        sleeps = []
        responses = [
            PermissionError("access denied"),
            SimpleNamespace(stdout=mediawiki_response(tabx(("id",), [["ok"]])), stderr=""),
        ]

        def runner(command, **kwargs):
            calls.append(command)
            result = responses.pop(0)
            if isinstance(result, Exception):
                raise result
            return result

        rows = stats.fetch_tabx_page(
            "Data:Example.tabx",
            ("id",),
            run_command=runner,
            sleep=sleeps.append,
            which=lambda name: "C:/curl.exe",
        )
        self.assertEqual(rows, [{"id": "ok"}])
        self.assertEqual(len(calls), 2)
        self.assertEqual(sleeps, [1])

    def test_oserror_exhausts_four_attempts_with_a_clear_error(self):
        calls = []

        def runner(command, **kwargs):
            calls.append(command)
            raise FileNotFoundError("curl became unavailable")

        with self.assertRaisesRegex(stats.FetchError, "after 4 attempts:.*curl became unavailable"):
            stats.fetch_tabx_page(
                "Data:Example.tabx",
                ("id",),
                run_command=runner,
                sleep=lambda seconds: None,
                which=lambda name: "C:/curl.exe",
            )
        self.assertEqual(len(calls), 4)

    def test_retries_invalid_nested_tabx_json_then_succeeds(self):
        calls = []
        sleeps = []
        responses = [
            SimpleNamespace(stdout=mediawiki_response("{broken"), stderr=""),
            SimpleNamespace(stdout=mediawiki_response(tabx(("id",), [["ok"]])), stderr=""),
        ]

        def runner(command, **kwargs):
            calls.append(command)
            return responses.pop(0)

        rows = stats.fetch_tabx_page(
            "Data:Example.tabx",
            ("id",),
            run_command=runner,
            sleep=sleeps.append,
            which=lambda name: "C:/curl.exe",
        )
        self.assertEqual(rows, [{"id": "ok"}])
        self.assertEqual(len(calls), 2)
        self.assertEqual(sleeps, [1])

    def test_mediawiki_error_missing_and_structural_responses_are_not_retried(self):
        responses = [
            ("error", json.dumps({"error": {"code": "badtitle"}}), "missing query"),
            ("missing", json.dumps({}), "missing query"),
            ("structural", json.dumps({"query": {"pages": []}}), "invalid pages list"),
        ]
        for label, response, message in responses:
            with self.subTest(label=label):
                calls = []

                def runner(command, **kwargs):
                    calls.append(command)
                    return SimpleNamespace(stdout=response, stderr="")

                with self.assertRaisesRegex(stats.FetchError, message):
                    stats.fetch_tabx_page(
                        "Data:Example.tabx",
                        ("id",),
                        run_command=runner,
                        sleep=lambda seconds: None,
                        which=lambda name: "C:/curl.exe",
                    )
                self.assertEqual(len(calls), 1)


class WorkbookAndCliTests(unittest.TestCase):
    def test_workbook_has_two_structured_sides_and_expected_presentation(self):
        rows = {
            "进攻方": [
                stats.OperatorRow(
                    name="Alice",
                    speed=3,
                    order=1,
                    primary_automatic=(
                        stats.WeaponRate("Fast", 900, 2),
                        stats.WeaponRate("Slow", 700, 1),
                    ),
                    secondary_automatic=(stats.WeaponRate("Side Auto", 1000, 3),),
                    primary_semiautomatic=(stats.WeaponRate("DMR", 0, 4),),
                    secondary_shotguns=(stats.WeaponRate("Shorty", 0, 5),),
                    secondary_gadgets=("手雷×2", "烟雾弹×2"),
                )
            ],
            "防守方": [
                stats.OperatorRow(
                    name="Bob",
                    speed=2,
                    order=2,
                    primary_automatic=(),
                    secondary_automatic=(),
                    primary_semiautomatic=(),
                    secondary_shotguns=(),
                    secondary_gadgets=("倒刺铁丝网×2",),
                )
            ],
        }
        with tempfile.TemporaryDirectory() as directory:
            icon_dir = Path(directory) / "icons"
            badge_dir = icon_dir / "badge"
            badge_dir.mkdir(parents=True)
            for name, color in (("Alice", "red"), ("Bob", "blue")):
                Image.new("RGBA", (32, 32), color).save(badge_dir / (stats.operator_key(name) + ".png"))
            ratings = {
                "alice": stats.OperatorRating("S", 100),
                "bob": stats.OperatorRating("boof", 0),
            }
            output = Path(directory) / "stats.xlsx"
            stats.write_workbook(
                output,
                rows,
                ratings,
                icon_dir,
                make_report_sources(with_changes=True),
            )
            workbook = load_workbook(output)

        self.assertEqual(workbook.sheetnames, ["进攻方", "防守方", "补丁说明"])
        attackers = workbook["进攻方"]
        defenders = workbook["防守方"]
        for sheet in (attackers, defenders):
            self.assertEqual(
                [cell.value for cell in sheet[1]],
                [
                    "图标", "干员", "速度", "主手自动枪械（射速，发/分钟）",
                    "副手自动枪械（射速，发/分钟）", "主狙",
                    "副手霰弹", "次要装备", "Athieno评分",
                ],
            )
            self.assertTrue(sheet["A1"].font.bold)
            self.assertEqual(sheet["A1"].fill.fill_type, "solid")
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertIsNone(sheet.auto_filter.ref)
            self.assertTrue(sheet.tables)
            self.assertEqual(next(iter(sheet.tables.values())).ref, "A1:I2")
            self.assertIn(
                "补丁区间：", sheet.cell(sheet.max_row, 1).value
            )
            self.assertIn(str(sheet.max_row), str(sheet.print_area))
            for column in "DEFGH":
                self.assertTrue(sheet[f"{column}2"].alignment.wrap_text)
            self.assertEqual(len(sheet._images), 1)
            self.assertGreater(sheet.row_dimensions[2].height, 30)
            self.assertEqual(
                tuple(sheet.column_dimensions[column].width for column in "ABCDEFGHI"),
                (8.0, 18.0, 10.0, 34.0, 32.0, 24.0, 20.0, 32.0, 14.0),
            )
        self.assertEqual(attackers["C2"].value, 3)
        self.assertIsInstance(attackers["C2"].value, (int, float))
        self.assertEqual(attackers["D2"].value, "Fast（900）\nSlow（700）")
        self.assertEqual(attackers["E2"].value, "Side Auto（1000）")
        self.assertEqual(attackers["F2"].value, "DMR")
        self.assertEqual(attackers["G2"].value, "Shorty")
        self.assertEqual(attackers["H2"].value, "手雷×2\n烟雾弹×2")
        self.assertEqual(attackers["I2"].value, 100)
        self.assertEqual(attackers["I2"].fill.fgColor.rgb, "00E74C3C")
        self.assertEqual(defenders["D2"].value, "无自动枪械")
        self.assertEqual(defenders["E2"].value, "无自动枪械")
        self.assertEqual(defenders["F2"].value, "无")
        self.assertEqual(defenders["G2"].value, "无")
        self.assertEqual(defenders["I2"].fill.fgColor.rgb, "007F8C8D")
        self.assertIsNotNone(attackers["I1"].comment)
        self.assertEqual(
            workbook["补丁说明"]["A1"].value,
            "Y11S2 视频评分后续补丁说明",
        )

    def test_main_prints_counts_and_resolved_output_path(self):
        rows = {
            "进攻方": [
                stats.OperatorRow(
                    "Alice", 3, 1, (stats.WeaponRate("Gun", 900, 1),), (), (), (), ("手雷×2",)
                )
            ],
            "防守方": [stats.OperatorRow("Bob", 2, 2, (), (), (), (), ("倒刺铁丝网×2",))],
        }
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "out.xlsx"
            stdout = io.StringIO()
            with contextlib.redirect_stdout(stdout):
                result = stats.main(
                    ["--output", str(output), "--icons-dir", str(Path(directory) / "icons")],
                    fetcher=lambda: rows,
                    rating_loader=lambda path, names: dict(
                        alice=stats.OperatorRating("S", 100),
                        bob=stats.OperatorRating("A", 85),
                    ),
                    icon_preparer=lambda rows, path: _make_test_icons(path, ("Alice", "Bob")),
                    source_loader=lambda path: make_report_sources(
                        with_changes=True
                    ),
                )

            self.assertEqual(result, 0)
            self.assertTrue(output.exists())
            summary = stdout.getvalue()
        self.assertIn("进攻方干员：1", summary)
        self.assertIn("防守方干员：1", summary)
        self.assertIn("唯一自动枪械：1", summary)
        self.assertIn(str(output.resolve()), summary)


def _make_test_icons(path, names):
    badge_dir = path / "badge"
    badge_dir.mkdir(parents=True, exist_ok=True)
    for name in names:
        Image.new("RGBA", (32, 32), "red").save(badge_dir / (stats.operator_key(name) + ".png"))
    return path


class RatingAndIconTests(unittest.TestCase):
    def test_operator_key_normalizes_accents_and_spaces(self):
        self.assertEqual(stats.operator_key("Capitão"), "capitao")
        self.assertEqual(stats.operator_key("Jäger"), "jager")
        self.assertEqual(stats.operator_key("Nøkk"), "nokk")
        self.assertEqual(stats.operator_key("Skopós"), "skopos")
        self.assertEqual(stats.operator_key("Solid Snake"), "solid-snake")

    def test_rating_file_covers_every_operator_once(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "ratings.json"
            path.write_text(json.dumps({
                "score_map": {"S": 100, "A": 85},
                "tiers": {"S": ["alice"], "A": ["bob"]},
            }), encoding="utf-8")
            ratings = stats.load_ratings(path, ("Alice", "Bob"))
            self.assertEqual(ratings["alice"], stats.OperatorRating("S", 100))
            self.assertEqual(ratings["bob"], stats.OperatorRating("A", 85))

            with self.assertRaisesRegex(stats.DataFormatError, "duplicate rating"):
                path.write_text(json.dumps({
                    "score_map": {"S": 100, "A": 85},
                    "tiers": {"S": ["alice"], "A": ["alice", "bob"]},
                }), encoding="utf-8")
                stats.load_ratings(path, ("Alice", "Bob"))

            with self.assertRaisesRegex(stats.DataFormatError, "rating coverage mismatch"):
                path.write_text(json.dumps({
                    "score_map": {"S": 100}, "tiers": {"S": ["alice"]}
                }), encoding="utf-8")
                stats.load_ratings(path, ("Alice", "Bob"))

    def test_production_rating_file_has_77_entries(self):
        path = (
            Path(__file__).resolve().parents[1]
            / "inputs"
            / "athieno"
            / "latest.json"
        )
        names = [name for tier in json.loads(path.read_text(encoding="utf-8"))["tiers"].values() for name in tier]
        self.assertEqual(len(names), 77)
        self.assertEqual(len(set(names)), 77)

    def test_builds_transparent_pure_white_icon(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.png"
            output = Path(directory) / "output.png"
            image = Image.new("RGBA", (20, 40), (0, 0, 0, 0))
            for x in range(5, 15):
                for y in range(10, 30):
                    image.putpixel((x, y), (20, 30, 40, 255))
            image.save(source)

            stats.make_white_operator_icon(source, output)
            rendered = Image.open(output).convert("RGBA")

        self.assertEqual(rendered.size, (96, 96))
        self.assertEqual(rendered.getpixel((0, 0)), (0, 0, 0, 0))
        self.assertEqual(rendered.getpixel((48, 48)), (255, 255, 255, 255))


class BatchLauncherTests(unittest.TestCase):
    def test_batch_launcher_runs_from_project_directory_and_preserves_exit_code(self):
        launcher = Path(__file__).resolve().parents[1] / "run_r6_report.bat"
        self.assertTrue(launcher.exists(), "batch launcher does not exist")

        content = launcher.read_text(encoding="utf-8")
        self.assertIn('cd /d "%~dp0"', content)
        self.assertIn('where python', content)
        self.assertIn('where py', content)
        self.assertIn("-m r6_report.operator_stats", content)
        self.assertIn('--inputs-dir "%~dp0inputs"', content)
        self.assertIn(
            '--output "%~dp0~temp\\r6_operator_stats.xlsx"',
            content,
        )
        self.assertIn('set "EXIT_CODE=%ERRORLEVEL%"', content)
        self.assertIn('exit /b %EXIT_CODE%', content)


if __name__ == "__main__":
    unittest.main()
