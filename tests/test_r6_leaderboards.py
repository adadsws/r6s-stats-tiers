import contextlib
import io
import tempfile
import unittest
from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader
from reportlab.lib.pagesizes import A2
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

import _path_setup
from r6_report import leaderboards as lb
from r6_report import pdf_leaderboards as pdf_lb
from r6_report import report_theme as theme
from r6_report import tier_chart as tier
from r6_report.sources import PatchChange
from source_fixtures import make_report_sources


def make_card(
    name,
    order,
    *,
    side="进攻方",
    visible_tier="A",
    score=85,
    speed=2,
    primary=(800,),
    secondary=(),
    has_semiautomatic=False,
    has_secondary_shotgun=False,
    gadgets=("烟雾弹",),
):
    primary_weapons = tuple(
        tier.WeaponItem(
            "Primary %d" % index,
            "primary-%d" % index,
            rate,
        )
        for index, rate in enumerate(primary, start=1)
    )
    secondary_weapons = tuple(
        tier.WeaponItem(
            "Secondary %d" % index,
            "secondary-%d" % index,
            rate,
        )
        for index, rate in enumerate(secondary, start=1)
    )
    return tier.OperatorCard(
        side=side,
        name=name,
        speed=speed,
        score=score,
        tier=visible_tier,
        primary_rpms=primary,
        secondary_rpms=secondary,
        has_semiautomatic=has_semiautomatic,
        has_secondary_shotgun=has_secondary_shotgun,
        gadgets=tuple(tier.GadgetItem(item, 2) for item in gadgets),
        source_order=order,
        primary_weapons=primary_weapons,
        secondary_weapons=secondary_weapons,
        semiautomatic_weapons=(
            (tier.WeaponItem("Primary Marksman", "primary-marksman"),)
            if has_semiautomatic
            else ()
        ),
        secondary_shotguns=(
            (tier.WeaponItem("Secondary Shotgun", "secondary-shotgun"),)
            if has_secondary_shotgun
            else ()
        ),
    )


class LeaderboardClassificationTests(unittest.TestCase):
    def test_loads_each_required_weapon_icon_once(self):
        items = (
            tier.WeaponItem("R4-C", "r4-c", 860),
            tier.WeaponItem("R4-C", "r4-c", 860),
            tier.WeaponItem("G8A1", "g8a1", 850),
        )
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            Image.new("RGBA", (24, 8), "black").save(root / "r4-c.png")
            Image.new("RGBA", (24, 8), "black").save(root / "g8a1.png")

            paths = lb.load_weapon_icons(items, root)

        self.assertEqual(tuple(paths), ("r4-c", "g8a1"))

    def test_defines_five_dimensions_in_global_priority_order(self):
        self.assertEqual(
            lb.DIMENSION_ORDER,
            ("video", "primary_rpm", "speed", "rare", "gadget"),
        )
        self.assertEqual(tuple(lb.LEADERBOARD_SPECS), lb.DIMENSION_ORDER)

    def test_primary_rpm_uses_fixed_four_tiers(self):
        cases = (
            ((1200,), "Ⅰ"),
            ((860,), "Ⅰ"),
            ((859,), "Ⅱ"),
            ((780,), "Ⅱ"),
            ((779,), "Ⅲ"),
            ((700,), "Ⅲ"),
            ((699,), "Ⅳ"),
            ((), "Ⅳ"),
        )
        for primary, expected in cases:
            with self.subTest(primary=primary):
                card = make_card("Boundary", 1, primary=primary)
                self.assertEqual(
                    lb.bands_for_card(card, "primary_rpm", "进攻方"),
                    (expected,),
                )

    def test_rare_weapon_memberships_repeat_and_fallback_is_exclusive(self):
        triple = make_card(
            "Triple",
            1,
            primary=(900,),
            secondary=(1270,),
            has_semiautomatic=True,
            has_secondary_shotgun=True,
        )
        none = make_card(
            "None",
            2,
            primary=(900,),
            secondary=(),
            has_semiautomatic=False,
            has_secondary_shotgun=False,
        )

        self.assertEqual(
            lb.bands_for_card(triple, "rare", "进攻方"),
            ("副喷", "主狙", "副自"),
        )
        self.assertEqual(
            lb.bands_for_card(none, "rare", "进攻方"),
            ("都无",),
        )

    def test_gadget_memberships_repeat_by_side_and_fallback_is_exclusive(self):
        attacker = make_card(
            "Attacker",
            1,
            gadgets=("破片手榴弹", "闪光弹", "硬突破炸药", "烟雾弹"),
        )
        defender = make_card(
            "Defender",
            2,
            side="防守方",
            gadgets=("遥控炸药", "机动护盾", "冲击手榴弹", "倒刺铁丝网"),
        )
        no_target = make_card(
            "No Target",
            3,
            gadgets=("烟雾弹", "阔剑地雷"),
        )

        self.assertEqual(
            lb.bands_for_card(attacker, "gadget", "进攻方"),
            ("手雷", "眩晕手榴弹", "硬突破炸药"),
        )
        self.assertEqual(
            lb.bands_for_card(defender, "gadget", "防守方"),
            ("遥控炸药", "机动护盾", "冲击手榴弹"),
        )
        self.assertEqual(
            lb.bands_for_card(no_target, "gadget", "进攻方"),
            ("这些都无",),
        )

    def test_video_and_speed_have_single_memberships(self):
        card = make_card(
            "Single",
            1,
            visible_tier="S",
            speed=3,
        )

        self.assertEqual(
            lb.bands_for_card(card, "video", "进攻方"),
            ("S",),
        )
        self.assertEqual(
            lb.bands_for_card(card, "speed", "进攻方"),
            ("3速",),
        )

    def test_collapses_post_video_patch_directions_into_card_markers(self):
        cases = (
            ((), ""),
            (("增强",), "+"),
            (("增强", "增强"), "+"),
            (("削弱",), "-"),
            (("削弱", "削弱"), "-"),
            (("增强", "削弱"), "~"),
            (("混合",), "~"),
        )
        for directions, expected in cases:
            with self.subTest(directions=directions):
                self.assertEqual(
                    lb.patch_direction_marker(directions),
                    expected,
                )

        self.assertEqual(
            lb.patch_markers(make_report_sources(with_changes=True)),
            {"Alice": "~", "Bob": "-"},
        )

    def test_excludes_only_pure_gadget_balance_from_operator_markers(self):
        cases = (
            (
                PatchChange(
                    "增强",
                    "Mute",
                    "防弹摄像头电磁脉冲波的爆炸范围提升。",
                ),
                False,
            ),
            (
                PatchChange("增强", "Wamai", "新增机动护盾。"),
                True,
            ),
            (
                PatchChange(
                    "混合",
                    "Wamai",
                    "冲击手榴弹被替换为遥控炸药。",
                ),
                True,
            ),
            (
                PatchChange(
                    "增强",
                    "Wamai",
                    "磁力销毁系统充能提高；新增机动护盾。",
                ),
                True,
            ),
            (
                PatchChange("削弱", "Ace", "技能持续时间缩短。"),
                True,
            ),
        )
        for change, expected in cases:
            with self.subTest(detail=change.detail):
                self.assertEqual(
                    lb.counts_as_operator_adjustment(change),
                    expected,
                )

        sources = make_report_sources()
        patch = replace(
            sources.patches[0],
            changes=(
                PatchChange(
                    "增强",
                    "Mute",
                    "防弹摄像头电磁脉冲波的爆炸范围提升。",
                ),
                PatchChange("削弱", "Ace", "技能持续时间缩短。"),
            ),
        )
        sources = replace(sources, patches=(patch,))

        self.assertEqual(lb.patch_markers(sources), {"Ace": "-"})


class LeaderboardSortingTests(unittest.TestCase):
    def test_primary_band_uses_raw_rpm_before_video(self):
        cards = [
            make_card(
                "Higher Video Lower RPM",
                1,
                visible_tier="S",
                primary=(860,),
            ),
            make_card(
                "Lower Video Higher RPM",
                2,
                visible_tier="A",
                primary=(900,),
            ),
        ]

        sorted_cards = lb.sort_cards_for_band(
            cards,
            "primary_rpm",
            "进攻方",
        )

        self.assertEqual(
            [card.name for card in sorted_cards],
            ["Lower Video Higher RPM", "Higher Video Lower RPM"],
        )

    def test_video_band_uses_raw_score_before_primary_rpm(self):
        cards = [
            make_card(
                "Lower Score Higher RPM",
                1,
                visible_tier="S",
                score=90,
                primary=(900,),
            ),
            make_card(
                "Higher Score Lower RPM",
                2,
                visible_tier="S",
                score=95,
                primary=(700,),
            ),
        ]

        sorted_cards = lb.sort_cards_for_band(cards, "video", "进攻方")

        self.assertEqual(
            [card.name for card in sorted_cards],
            ["Higher Score Lower RPM", "Lower Score Higher RPM"],
        )

    def test_video_band_uses_primary_then_speed_before_later_dimensions(self):
        cards = [
            make_card(
                "Lower Primary",
                1,
                visible_tier="S",
                primary=(800,),
                speed=3,
                has_secondary_shotgun=True,
                gadgets=("破片手榴弹",),
            ),
            make_card(
                "Higher Primary",
                2,
                visible_tier="S",
                primary=(900,),
                speed=1,
                gadgets=("烟雾弹",),
            ),
            make_card(
                "Same Primary Faster",
                3,
                visible_tier="S",
                primary=(800,),
                speed=2,
                has_semiautomatic=True,
                gadgets=("烟雾弹",),
            ),
        ]

        sorted_cards = lb.sort_cards_for_band(cards, "video", "进攻方")

        self.assertEqual(
            [card.name for card in sorted_cards],
            ["Higher Primary", "Lower Primary", "Same Primary Faster"],
        )

    def test_primary_band_uses_video_then_speed(self):
        cards = [
            make_card(
                "A Three",
                1,
                visible_tier="A",
                speed=3,
                primary=(900,),
            ),
            make_card(
                "S One",
                2,
                visible_tier="S",
                speed=1,
                primary=(900,),
            ),
            make_card(
                "A Two",
                3,
                visible_tier="A",
                speed=2,
                primary=(900,),
            ),
        ]

        sorted_cards = lb.sort_cards_for_band(
            cards,
            "primary_rpm",
            "进攻方",
        )

        self.assertEqual(
            [card.name for card in sorted_cards],
            ["S One", "A Three", "A Two"],
        )

    def test_speed_band_uses_video_then_primary(self):
        cards = [
            make_card(
                "A High Primary",
                1,
                visible_tier="A",
                speed=3,
                primary=(900,),
            ),
            make_card(
                "S Low Primary",
                2,
                visible_tier="S",
                speed=3,
                primary=(700,),
            ),
            make_card(
                "A Low Primary",
                3,
                visible_tier="A",
                speed=3,
                primary=(700,),
            ),
        ]

        sorted_cards = lb.sort_cards_for_band(cards, "speed", "进攻方")

        self.assertEqual(
            [card.name for card in sorted_cards],
            ["S Low Primary", "A High Primary", "A Low Primary"],
        )

    def test_rare_band_does_not_promote_other_rare_memberships(self):
        cards = [
            make_card(
                "Higher Video Current Membership Only",
                1,
                visible_tier="S",
                has_semiautomatic=True,
            ),
            make_card(
                "Lower Video Also Higher Membership",
                2,
                visible_tier="A",
                has_semiautomatic=True,
                has_secondary_shotgun=True,
            ),
        ]

        sorted_cards = lb.sort_cards_for_band(cards, "rare", "进攻方")

        self.assertEqual(
            [card.name for card in sorted_cards],
            [
                "Higher Video Current Membership Only",
                "Lower Video Also Higher Membership",
            ],
        )

    def test_gadget_band_does_not_promote_other_gadget_memberships(self):
        cards = [
            make_card(
                "Higher Video Current Membership Only",
                1,
                visible_tier="S",
                gadgets=("闪光弹",),
            ),
            make_card(
                "Lower Video Also Higher Membership",
                2,
                visible_tier="A",
                gadgets=("破片手榴弹", "闪光弹"),
            ),
        ]

        sorted_cards = lb.sort_cards_for_band(cards, "gadget", "进攻方")

        self.assertEqual(
            [card.name for card in sorted_cards],
            [
                "Higher Video Current Membership Only",
                "Lower Video Also Higher Membership",
            ],
        )

    def test_source_order_is_the_final_tie_breaker(self):
        cards = [
            make_card("Later", 4),
            make_card("Earlier", 2),
        ]

        sorted_cards = lb.sort_cards_for_band(cards, "video", "进攻方")

        self.assertEqual(
            [card.name for card in sorted_cards],
            ["Earlier", "Later"],
        )

    def test_grouping_repeats_only_rare_and_gadget_cards(self):
        repeated = make_card(
            "Repeated",
            1,
            secondary=(1270,),
            has_semiautomatic=True,
            has_secondary_shotgun=True,
            gadgets=("破片手榴弹", "闪光弹"),
        )
        ordinary = make_card("Ordinary", 2, gadgets=("烟雾弹",))
        cards = (repeated, ordinary)

        rare = lb.group_cards(cards, "rare", "进攻方")
        gadget = lb.group_cards(cards, "gadget", "进攻方")
        video = lb.group_cards(cards, "video", "进攻方")

        self.assertEqual(
            sum(card.name == "Repeated" for group in rare.values() for card in group),
            3,
        )
        self.assertEqual(
            sum(card.name == "Repeated" for group in gadget.values() for card in group),
            2,
        )
        self.assertEqual(
            sum(card.name == "Repeated" for group in video.values() for card in group),
            1,
        )
        self.assertEqual(
            {
                card.name
                for group in rare.values()
                for card in group
            },
            {"Repeated", "Ordinary"},
        )


class LeaderboardWorkbookTests(unittest.TestCase):
    @staticmethod
    def _card_tier_cell(sheet, tier_text):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.column > 1 and cell.value == tier_text:
                    return cell
        raise AssertionError("operator card tier not found: %s" % tier_text)

    def _write_workbook(
        self,
        root,
        dimension,
        attackers,
        defenders,
        report_sources=None,
    ):
        root.mkdir(parents=True, exist_ok=True)
        badge_dir = root / "badges"
        badge_dir.mkdir(exist_ok=True)
        all_cards = tuple(attackers) + tuple(defenders)
        for card in all_cards:
            Image.new("RGBA", (64, 64), "blue").save(
                badge_dir / (tier.operator_key(card.name) + ".png")
            )

        gadget_names = {
            gadget.name
            for card in all_cards
            for gadget in card.gadgets
        }
        gadget_icons = {}
        for index, name in enumerate(sorted(gadget_names)):
            icon = root / ("gadget-%d.png" % index)
            Image.new("RGBA", (32, 32), "black").save(icon)
            gadget_icons[name] = icon

        weapon_icons = {}
        weapon_items = (
            weapon
            for card in all_cards
            for group in (
                card.primary_weapons,
                card.secondary_weapons,
                card.semiautomatic_weapons,
                card.secondary_shotguns,
            )
            for weapon in group
        )
        for weapon in weapon_items:
            if weapon.icon_key in weapon_icons:
                continue
            icon = root / ("weapon-%s.png" % weapon.icon_key)
            Image.new("RGBA", (36, 12), "black").save(icon)
            weapon_icons[weapon.icon_key] = icon

        output = root / ("%s.xlsx" % dimension)
        lb.write_leaderboard_workbook(
            output,
            lb.LEADERBOARD_SPECS[dimension],
            {"进攻方": attackers, "防守方": defenders},
            badge_dir,
            gadget_icons,
            weapon_icons,
            report_sources or make_report_sources(),
        )
        return load_workbook(output)

    def test_renders_operator_name_below_full_size_badge(self):
        attacker = replace(
            make_card(
                "Multi Weapon",
                1,
                primary=(980, 860, 750),
                secondary=(1270, 1000),
                has_semiautomatic=True,
                has_secondary_shotgun=True,
            ),
            semiautomatic_weapons=(
                tier.WeaponItem("DMR One", "dmr-one"),
                tier.WeaponItem("DMR Two", "dmr-two"),
            ),
            secondary_shotguns=(
                tier.WeaponItem("Shotgun One", "shotgun-one"),
                tier.WeaponItem("Shotgun Two", "shotgun-two"),
            ),
        )

        with tempfile.TemporaryDirectory() as directory:
            workbook = self._write_workbook(
                Path(directory),
                "video",
                (attacker,),
                (),
            )

        sheet = workbook["进攻方视频Tier榜"]
        tier_cell = self._card_tier_cell(sheet, "A")
        card_row = tier_cell.row
        badge_column = tier_cell.column - 1
        primary_column = badge_column + 1
        speed_column = badge_column + 3
        secondary_column = badge_column + 4
        badge_cell = sheet.cell(card_row, badge_column)
        self.assertEqual(badge_cell.value, "Multi Weapon")
        self.assertEqual(badge_cell.alignment.vertical, "bottom")
        self.assertTrue(badge_cell.alignment.shrink_to_fit)
        badge_image = next(
            image
            for image in sheet._images
            if image.anchor._from.row == card_row - 1
            and image.anchor._from.col == badge_column - 1
        )
        self.assertEqual(round(badge_image.anchor.ext.cx / 9525), 48)
        self.assertEqual(round(badge_image.anchor.ext.cy / 9525), 48)
        self.assertEqual(
            [
                sheet.cell(card_row, primary_column).value,
                sheet.cell(card_row, speed_column).value,
                sheet.cell(card_row, secondary_column).value,
            ],
            ["A", "2速", "主狙 ✓"],
        )
        self.assertEqual(
            [sheet.cell(card_row + offset, primary_column).value for offset in (1, 2, 3)],
            ["主 980", "主 860", "主 750"],
        )
        self.assertEqual(
            [sheet.cell(card_row + offset, secondary_column).value for offset in (1, 2, 3)],
            ["副喷 ✓", "副 1270", "副 1000"],
        )
        weapon_images = [
            image
            for image in sheet._images
            if image.anchor._from.row
            in tuple(card_row - 1 + offset for offset in range(4))
            and round(image.anchor.ext.cy / 9525)
            <= theme.XLSX_WEAPON_ICON_HEIGHT_PX
        ]
        self.assertEqual(
            [
                sum(
                    image.anchor._from.row == card_row + offset - 1
                    and image.anchor._from.col == primary_column - 1
                    for image in weapon_images
                )
                for offset in (1, 2, 3)
            ],
            [1, 1, 1],
        )
        self.assertEqual(
            [
                sum(
                    image.anchor._from.row == card_row + offset - 1
                    and image.anchor._from.col == secondary_column - 1
                    for image in weapon_images
                )
                for offset in (1, 2, 3)
            ],
            [2, 1, 1],
        )

    def test_renders_five_row_cards_and_wraps_sixth_inside_same_band(self):
        attackers = tuple(
            make_card(
                "Attacker %d" % index,
                index,
                visible_tier="S",
                speed=3,
                primary=(900, 700),
                secondary=(1270,),
                has_semiautomatic=True,
                has_secondary_shotgun=True,
                gadgets=("破片手榴弹", "烟雾弹"),
            )
            for index in range(1, 7)
        )
        defender = make_card(
            "Defender",
            1,
            side="防守方",
            visible_tier="F",
            primary=(),
            gadgets=("倒刺铁丝网",),
        )

        with tempfile.TemporaryDirectory() as directory:
            workbook = self._write_workbook(
                Path(directory),
                "video",
                attackers,
                (defender,),
            )

        self.assertEqual(
            workbook.sheetnames,
            ["进攻方视频Tier榜", "防守方视频Tier榜", "补丁说明"],
        )
        sheet = workbook["进攻方视频Tier榜"]
        merged = {str(item) for item in sheet.merged_cells.ranges}
        self.assertIn("A3:A12", merged)
        self.assertIn("B3:B6", merged)
        self.assertIn("C3:D3", merged)
        self.assertIn("F3:H3", merged)
        self.assertIn("C4:E4", merged)
        self.assertIn("F4:H4", merged)
        self.assertIn("C5:E5", merged)
        self.assertIn("F5:H5", merged)
        self.assertIn("C6:E6", merged)
        self.assertIn("F6:H6", merged)
        self.assertEqual(sheet["C3"].value, "S")
        self.assertEqual(sheet["E3"].value, "3速")
        self.assertEqual(sheet["F3"].value, "主狙 ✓")
        self.assertEqual(sheet["C4"].value, "主 900")
        self.assertEqual(sheet["F4"].value, "副喷 ✓")
        self.assertEqual(sheet["C5"].value, "主 700")
        self.assertEqual(sheet["F5"].value, "副 1270")
        self.assertIsNone(sheet["C6"].value)
        self.assertIsNone(sheet["F6"].value)
        self.assertEqual(sheet["C3"].border.right.style, "thin")
        self.assertEqual(sheet["E3"].border.left.style, "thin")
        self.assertEqual(sheet["E3"].border.right.style, "medium")
        self.assertEqual(sheet["F3"].border.left.style, "medium")
        self.assertEqual(sheet["C4"].border.right.style, "medium")
        self.assertEqual(sheet["F4"].border.left.style, "medium")
        self.assertEqual(sheet["C4"].border.bottom.style, "thin")
        self.assertEqual(sheet["F4"].border.bottom.style, "medium")
        self.assertEqual(
            sheet["C4"].border.bottom.color.rgb[-6:],
            "8F969E",
        )
        self.assertEqual(sheet["B7"].border.right.style, "thin")
        self.assertEqual(sheet["B7"].border.bottom.style, "medium")
        self.assertAlmostEqual(
            sheet.column_dimensions["C"].width
            + sheet.column_dimensions["D"].width,
            sheet.column_dimensions["E"].width,
        )
        self.assertEqual(sheet["C8"].value, "S")
        self.assertEqual(sheet["E8"].value, "3速")
        self.assertEqual(sheet["B3"].value, "Attacker 1")
        self.assertEqual(sheet["B8"].value, "Attacker 6")
        self.assertEqual(sheet.freeze_panes, "B3")
        self.assertIn(
            "补丁区间：", sheet.cell(sheet.max_row, 1).value
        )
        self.assertIn(
            "A%d:AO%d" % (sheet.max_row, sheet.max_row),
            merged,
        )
        self.assertEqual(
            sheet["C3"].fill.fgColor.rgb[-6:],
            tier.TIER_COLORS["S"],
        )
        self.assertEqual(
            sheet["E3"].fill.fgColor.rgb[-6:],
            "F4F5F7",
        )
        missing_sheet = workbook["防守方视频Tier榜"]
        defender_tier = self._card_tier_cell(missing_sheet, "F")
        missing_cells = (
            missing_sheet.cell(defender_tier.row, defender_tier.column + 3),
            missing_sheet.cell(defender_tier.row + 1, defender_tier.column + 3),
        )
        self.assertEqual(
            tuple(cell.value for cell in missing_cells),
            ("主狙 -", "副喷 -"),
        )
        for cell in missing_cells:
            self.assertEqual(
                cell.fill.fgColor.rgb[-6:],
                theme.MISSING_FILL,
            )
        self.assertIsNone(workbook["补丁说明"].freeze_panes)

    def test_appends_patch_marker_after_video_tier_on_every_leaderboard(self):
        attacker = make_card("Alice", 1, visible_tier="A")
        defender = make_card(
            "Bob",
            1,
            side="防守方",
            visible_tier="B",
            gadgets=("遥控炸药",),
        )

        with tempfile.TemporaryDirectory() as directory:
            workbook = self._write_workbook(
                Path(directory),
                "video",
                (attacker,),
                (defender,),
                make_report_sources(with_changes=True),
            )

        attacker_tier = self._card_tier_cell(
            workbook["进攻方视频Tier榜"],
            "A~",
        )
        defender_tier = self._card_tier_cell(
            workbook["防守方视频Tier榜"],
            "B-",
        )
        self.assertEqual(attacker_tier.value, "A~")
        self.assertEqual(defender_tier.value, "B-")

    def test_card_text_is_not_highlighted_red(self):
        attacker = make_card(
            "Highlighted",
            1,
            visible_tier="S",
            speed=3,
            primary=(900,),
            secondary=(1270,),
            has_semiautomatic=True,
            has_secondary_shotgun=True,
            gadgets=("破片手榴弹",),
        )
        defender = make_card(
            "Defender",
            1,
            side="防守方",
            primary=(900,),
            gadgets=("遥控炸药",),
        )

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            primary = self._write_workbook(
                root / "primary",
                "primary_rpm",
                (attacker,),
                (defender,),
            )
            speed = self._write_workbook(
                root / "speed",
                "speed",
                (attacker,),
                (defender,),
            )
            rare = self._write_workbook(
                root / "rare",
                "rare",
                (attacker,),
                (defender,),
            )

        primary_sheet = primary["进攻方主武器射速榜"]
        speed_sheet = speed["进攻方速度榜"]
        rare_sheet = rare["进攻方稀有枪械榜"]
        self.assertNotEqual(primary_sheet["C4"].font.color.rgb[-6:], "E74C3C")
        self.assertFalse(primary_sheet["C4"].font.bold)
        self.assertNotEqual(speed_sheet["C4"].font.color.rgb[-6:], "E74C3C")
        self.assertFalse(speed_sheet["C4"].font.bold)
        self.assertNotEqual(rare_sheet["F4"].font.color.rgb[-6:], "E74C3C")
        self.assertFalse(rare_sheet["F4"].font.bold)

    def test_gadget_token_has_no_red_highlight_border(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "source.png"
            output = root / "output.png"
            Image.new("RGBA", (32, 32), "black").save(source)

            lb.draw_gadget_token(source, 2, output)

            with Image.open(output) as rendered:
                red_pixels = [
                    pixel
                    for pixel in rendered.convert(
                        "RGBA"
                    ).get_flattened_data()
                    if pixel[:3] == (231, 76, 60)
                ]
            self.assertEqual(red_pixels, [])

    def test_gadget_token_omits_quantity_and_matches_body_row(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "source.png"
            output_two = root / "output-two.png"
            output_three = root / "output-three.png"
            Image.new("RGBA", (32, 16), "black").save(source)

            lb.draw_gadget_token(source, 2, output_two)
            lb.draw_gadget_token(source, 3, output_three)

            with (
                Image.open(output_two) as rendered_two,
                Image.open(output_three) as rendered_three,
            ):
                self.assertEqual(rendered_two.size, (24, 22))
                self.assertEqual(
                    rendered_two.convert("RGBA").tobytes(),
                    rendered_three.convert("RGBA").tobytes(),
                )
                alpha_bounds = rendered_two.getchannel("A").getbbox()
                self.assertIsNotNone(alpha_bounds)
                self.assertEqual(
                    max(
                        alpha_bounds[2] - alpha_bounds[0],
                        alpha_bounds[3] - alpha_bounds[1],
                    ),
                    22,
                )

    def test_wraps_seven_gadgets_inside_striker_and_sentry_cards(self):
        attacker = make_card(
            "Striker",
            1,
            gadgets=(
                "烟雾弹",
                "电磁脉冲式冲击弹",
                "破片手榴弹",
                "爆破炸药",
                "硬突破炸药",
                "阔剑地雷",
                "闪光弹",
            ),
        )
        defender = make_card(
            "Sentry",
            1,
            side="防守方",
            gadgets=(
                "感应警报器",
                "冲击手榴弹",
                "倒刺铁丝网",
                "遥控炸药",
                "观测工具阻拦器",
                "机动护盾",
                "防弹摄像头",
            ),
        )

        with tempfile.TemporaryDirectory() as directory:
            workbook = self._write_workbook(
                Path(directory),
                "video",
                (attacker,),
                (defender,),
            )

        for sheet_name in ("进攻方视频Tier榜", "防守方视频Tier榜"):
            sheet = workbook[sheet_name]
            tier_cell = self._card_tier_cell(sheet, "A")
            badge_column = tier_cell.column - 1
            gadget_row = tier_cell.row + 4
            gadget_images = [
                image
                for image in sheet._images
                if image.anchor._from.row == gadget_row - 1
                and badge_column - 1 <= image.anchor._from.col
                < badge_column + 6
            ]
            self.assertEqual(len(gadget_images), 7)
            self.assertEqual(
                [image.anchor._from.col for image in gadget_images],
                list(range(badge_column - 1, badge_column + 6)),
            )
            self.assertEqual(
                sheet.row_dimensions[gadget_row].height,
                theme.XLSX_CARD_BODY_ROW_PT,
            )
            row_height_px = round(theme.XLSX_CARD_BODY_ROW_PT * 96 / 72)
            for image in gadget_images:
                top_px = round(image.anchor._from.rowOff / 9525)
                height_px = round(image.anchor.ext.cy / 9525)
                self.assertLessEqual(
                    top_px + height_px,
                    row_height_px,
                )

    def test_sparse_gadget_keeps_its_fixed_xlsx_slot(self):
        attacker = make_card(
            "Sparse",
            1,
            gadgets=("烟雾弹",),
        )

        with tempfile.TemporaryDirectory() as directory:
            workbook = self._write_workbook(
                Path(directory),
                "video",
                (attacker,),
                (),
            )

        sheet = workbook["进攻方视频Tier榜"]
        tier_cell = self._card_tier_cell(sheet, "A")
        badge_column = tier_cell.column - 1
        gadget_row = tier_cell.row + 4
        gadget_images = [
            image
            for image in sheet._images
            if image.anchor._from.row == gadget_row - 1
        ]
        self.assertEqual(len(gadget_images), 1)
        self.assertEqual(
            (
                gadget_images[0].anchor._from.col,
                round(gadget_images[0].anchor._from.colOff / 9525),
                round(gadget_images[0].anchor._from.rowOff / 9525),
            ),
            (badge_column + 5, 6, 0),
        )
        self.assertEqual(
            sheet.row_dimensions[gadget_row].height,
            theme.XLSX_CARD_BODY_ROW_PT,
        )


class LeaderboardCliTests(unittest.TestCase):
    def test_rasterizes_pdf_pages_without_changing_page_geometry(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "source.pdf"
            renderer = canvas.Canvas(str(source), pagesize=(144, 72))
            renderer.setFillColorRGB(1, 0, 0)
            renderer.rect(0, 0, 144, 72, fill=1, stroke=0)
            renderer.showPage()
            renderer.setPageSize((72, 144))
            renderer.setFillColorRGB(0, 0, 1)
            renderer.rect(0, 0, 72, 144, fill=1, stroke=0)
            renderer.save()

            paths = pdf_lb.write_pdf_pages_as_png(
                source,
                root / "图片版",
                dpi=144,
            )

            self.assertEqual(
                tuple(path.name for path in paths),
                ("第1页.png", "第2页.png"),
            )
            with Image.open(paths[0]) as first, Image.open(paths[1]) as second:
                self.assertEqual(first.size, (288, 144))
                self.assertEqual(second.size, (144, 288))
                self.assertEqual(
                    first.convert("RGB").getpixel((144, 72)),
                    (255, 0, 0),
                )
                self.assertEqual(
                    second.convert("RGB").getpixel((72, 144)),
                    (0, 0, 255),
                )

    def test_pdf_gadget_image_preserves_aspect_ratio(self):
        with tempfile.TemporaryDirectory() as directory:
            icon = Path(directory) / "wide.png"
            Image.new("RGBA", (40, 20), "black").save(icon)

            self.assertEqual(
                pdf_lb._fit_image_size(icon, 20, 20),
                (20, 10),
            )

    def test_pdf_gadget_row_omits_name_and_uses_equal_slots(self):
        card = make_card("Attacker", 1, primary=(), gadgets=("烟雾弹",))
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            badge_dir = root / "badges"
            badge_dir.mkdir()
            Image.new("RGBA", (64, 64), "blue").save(
                badge_dir / "attacker.png"
            )
            icon_path = root / "wide.png"
            padded_icon = Image.new("RGBA", (290, 100), (0, 0, 0, 0))
            padded_icon.paste(
                Image.new("RGBA", (40, 20), "black"),
                (100, 40),
            )
            padded_icon.save(icon_path)

            pdf_lb._register_fonts()
            card_table = pdf_lb._card_flowable(
                card,
                badge_dir,
                {"烟雾弹": icon_path},
                {},
                "",
            )

            gadget_table = card_table._cellvalues[1][0]
            self.assertEqual(len(gadget_table._cellvalues), 1)
            self.assertEqual(
                gadget_table._colWidths,
                [11 * mm] * 7,
            )
            self.assertEqual(
                gadget_table._cellvalues[0][:6],
                ["", "", "", "", "", ""],
            )
            gadget_icon = gadget_table._cellvalues[0][6]
            self.assertIsInstance(gadget_icon, pdf_lb.Image)
            self.assertEqual(gadget_icon.hAlign, "CENTER")
            self.assertAlmostEqual(
                gadget_icon.drawWidth,
                6 * mm,
                delta=0.01,
            )
            self.assertAlmostEqual(
                gadget_icon.drawHeight,
                3 * mm,
                delta=0.01,
            )

    def test_pdf_renders_every_weapon_icon_in_fixed_height_rows(self):
        card = replace(
            make_card(
                "Multi Weapon",
                1,
                primary=(980, 860, 750),
                secondary=(1270, 1000),
                has_semiautomatic=True,
                has_secondary_shotgun=True,
            ),
            semiautomatic_weapons=(
                tier.WeaponItem("DMR One", "dmr-one"),
                tier.WeaponItem("DMR Two", "dmr-two"),
            ),
            secondary_shotguns=(
                tier.WeaponItem("Shotgun One", "shotgun-one"),
                tier.WeaponItem("Shotgun Two", "shotgun-two"),
            ),
        )
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            badge_dir = root / "badges"
            badge_dir.mkdir()
            Image.new("RGBA", (64, 64), "blue").save(
                badge_dir / "multi_weapon.png"
            )
            gadget_icon = root / "gadget.png"
            Image.new("RGBA", (32, 32), "black").save(gadget_icon)
            weapon_icons = {}
            for weapon in lb.card_weapon_items({"进攻方": (card,), "防守方": ()}):
                icon = root / (weapon.icon_key + ".png")
                Image.new("RGBA", (36, 12), "black").save(icon)
                weapon_icons[weapon.icon_key] = icon

            pdf_lb._register_fonts()
            card_table = pdf_lb._card_flowable(
                card,
                badge_dir,
                {"烟雾弹": gadget_icon},
                weapon_icons,
                "",
            )

        details = card_table._cellvalues[0][1]
        badge_panel = card_table._cellvalues[0][0]
        self.assertIsInstance(badge_panel, pdf_lb.Table)
        badge_image = badge_panel._cellvalues[0][0]
        badge_name = badge_panel._cellvalues[1][0]
        self.assertIsInstance(badge_image, pdf_lb.Image)
        self.assertAlmostEqual(badge_image._width, 15 * mm)
        self.assertAlmostEqual(badge_image._height, 15 * mm)
        self.assertEqual(badge_name.getPlainText(), "Multi Weapon")

        def image_count(value):
            if isinstance(value, pdf_lb.Image):
                return 1
            if isinstance(value, pdf_lb.Table):
                return sum(
                    image_count(cell)
                    for row in value._cellvalues
                    for cell in row
                )
            return 0

        self.assertEqual(len(details._cellvalues), 4)
        summary = details._cellvalues[0][0]
        self.assertIsInstance(summary, pdf_lb.Table)
        self.assertEqual(summary._cellvalues[0][0].getPlainText(), "A")
        self.assertEqual(summary._cellvalues[0][1].getPlainText(), "2速")
        self.assertEqual(summary._colWidths[0], summary._colWidths[1])
        self.assertTrue(
            any(
                command[0] == "BACKGROUND"
                and command[1:3] == ((0, 0), (0, 0))
                and command[3].hexval() == "0xf39c12"
                for command in summary._bkgrndcmds
            )
        )
        self.assertEqual(
            details._cellvalues[0][3]._cellvalues[0][0]
            ._content[0]
            .getPlainText(),
            "主狙 ✓",
        )
        self.assertIn(("SPAN", (0, 0), (2, 0)), details._spanCmds)
        self.assertIn(("SPAN", (3, 0), (5, 0)), details._spanCmds)
        detail_grid = next(
            command
            for command in details._linecmds
            if command[0] == "GRID"
        )
        self.assertEqual(detail_grid[3], 0.55)
        self.assertEqual(detail_grid[4].hexval(), "0x8f969e")
        self.assertTrue(
            any(
                command[:4] == ("LINEBEFORE", (3, 0), (3, -1), 0.85)
                for command in details._linecmds
            )
        )
        self.assertTrue(
            any(
                command[:4] == ("LINEBELOW", (3, 1), (5, 1), 0.85)
                for command in details._linecmds
            )
        )
        self.assertEqual(image_count(details._cellvalues[0][3]), 2)
        self.assertEqual(
            [image_count(details._cellvalues[row][0]) for row in range(1, 4)],
            [1, 1, 1],
        )
        self.assertEqual(
            [
                details._cellvalues[row][0]._cellvalues[0][0]
                ._content[0]
                .getPlainText()
                for row in range(1, 4)
            ],
            ["主 980", "主 860", "主 750"],
        )
        self.assertEqual(
            [image_count(details._cellvalues[row][3]) for row in range(1, 4)],
            [2, 1, 1],
        )
        self.assertEqual(
            [
                details._cellvalues[row][3]._cellvalues[0][0]
                ._content[0]
                .getPlainText()
                for row in (2, 3)
            ],
            ["副 1270", "副 1000"],
        )
        primary_rpm_field = details._cellvalues[1][0]
        text_box = primary_rpm_field._cellvalues[0][0]
        _, unscaled_height = text_box._content[0].wrap(
            primary_rpm_field._colWidths[0],
            1000,
        )
        self.assertLessEqual(
            unscaled_height,
            text_box._content[0].style.leading,
        )
        text_box.canv = canvas.Canvas(io.BytesIO())
        _, wrapped_height = text_box.wrap(
            primary_rpm_field._colWidths[0],
            primary_rpm_field._rowHeights[0],
        )
        self.assertLessEqual(
            wrapped_height,
            primary_rpm_field._rowHeights[0],
        )
        self.assertEqual(
            details._rowHeights,
            [
                7 * mm,
                theme.PDF_CARD_BODY_ROW_MM * mm,
                theme.PDF_CARD_BODY_ROW_MM * mm,
                theme.PDF_CARD_BODY_ROW_MM * mm,
            ],
        )

    def test_pdf_patch_body_uses_white_text_only_in_direction_cell(self):
        self.assertEqual(
            pdf_lb._patch_text_color(1, 0, "增强"),
            "#" + theme.COLOURS["white"],
        )
        self.assertEqual(
            pdf_lb._patch_text_color(1, 1, "增强"),
            "#" + theme.COLOURS["text"],
        )

    def test_main_generates_workbooks_pdfs_and_unchanged_png_pages(self):
        attacker = make_card(
            "Attacker",
            1,
            visible_tier="S",
            speed=3,
            primary=(900,),
            secondary=(1270,),
            has_secondary_shotgun=True,
            gadgets=("破片手榴弹",),
        )
        defender = make_card(
            "Defender",
            1,
            side="防守方",
            visible_tier="F",
            speed=1,
            primary=(),
            has_semiautomatic=True,
            gadgets=("遥控炸药",),
        )
        cards = {"进攻方": [attacker], "防守方": [defender]}

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "stats.xlsx"
            source.touch()
            output_dir = root / "output"
            badge_dir = root / "badges"
            badge_dir.mkdir()
            for card in (attacker, defender):
                Image.new("RGBA", (64, 64), "blue").save(
                    badge_dir / (tier.operator_key(card.name) + ".png")
                )
            frag = root / "frag.png"
            nitro = root / "nitro.png"
            Image.new("RGBA", (32, 32), "black").save(frag)
            Image.new("RGBA", (32, 32), "black").save(nitro)
            gadget_icons = {
                "破片手榴弹": frag,
                "遥控炸药": nitro,
            }
            weapon_icons = {}
            for weapon in lb.card_weapon_items(cards):
                icon = root / ("weapon-%s.png" % weapon.icon_key)
                Image.new("RGBA", (36, 12), "black").save(icon)
                weapon_icons[weapon.icon_key] = icon
            stdout = io.StringIO()
            stderr = io.StringIO()
            with contextlib.redirect_stdout(stdout), contextlib.redirect_stderr(stderr):
                result = lb.main(
                    [
                        "--input",
                        str(source),
                        "--output-dir",
                        str(output_dir),
                        "--icons-dir",
                        str(badge_dir),
                        "--gadget-icons-dir",
                        str(root / "gadget-cache"),
                    ],
                    card_loader=lambda path: cards,
                    gadget_icon_preparer=lambda items, path: gadget_icons,
                    weapon_icon_loader=lambda items, path: weapon_icons,
                    source_loader=lambda path: make_report_sources(),
                )

            self.assertEqual(result, 0)
            self.assertEqual(stderr.getvalue(), "")
            self.assertIn("进攻方干员：1", stdout.getvalue())
            self.assertIn("防守方干员：1", stdout.getvalue())
            self.assertEqual(
                tuple(path.name for path in sorted(output_dir.glob("*.xlsx"))),
                tuple(sorted(lb.EXPECTED_OUTPUTS)),
            )
            self.assertEqual(
                tuple(path.name for path in sorted(output_dir.glob("*.pdf"))),
                tuple(
                    sorted(
                        Path(filename).with_suffix(".pdf").name
                        for filename in lb.EXPECTED_OUTPUTS
                    )
                ),
            )
            self.assertEqual(
                len(tuple((output_dir / "图片版").glob("*/*.png"))),
                15,
            )
            for spec in lb.LEADERBOARD_SPECS.values():
                path = output_dir / spec.filename
                self.assertIn(str(path.resolve()), stdout.getvalue())
                workbook = load_workbook(path, read_only=True)
                self.assertEqual(
                    workbook.sheetnames,
                    [
                        "进攻方" + spec.sheet_suffix,
                        "防守方" + spec.sheet_suffix,
                        "补丁说明",
                    ],
                )
                workbook.close()
                pdf_path = path.with_suffix(".pdf")
                self.assertIn(str(pdf_path.resolve()), stdout.getvalue())
                reader = PdfReader(pdf_path)
                self.assertEqual(len(reader.pages), 3)
                png_dir = output_dir / "图片版" / pdf_path.stem
                png_paths = tuple(sorted(png_dir.glob("*.png")))
                self.assertEqual(
                    tuple(path.name for path in png_paths),
                    ("第1页.png", "第2页.png", "第3页.png"),
                )
                page_heights = []
                for page_number, page in enumerate(reader.pages, start=1):
                    self.assertAlmostEqual(
                        float(page.mediabox.width),
                        A2[0],
                        delta=0.1,
                    )
                    page_height = float(page.mediabox.height)
                    page_heights.append(page_height)
                    self.assertGreater(page_height, 100 * mm)
                    self.assertLess(page_height, A2[1] - 20 * mm)
                    self.assertIn(
                        "第 %d 页" % page_number,
                        page.extract_text() or "",
                    )
                    with Image.open(png_paths[page_number - 1]) as rendered:
                        expected_size = (
                            float(page.mediabox.width) * 2,
                            float(page.mediabox.height) * 2,
                        )
                        self.assertLessEqual(
                            abs(rendered.width - expected_size[0]),
                            1,
                        )
                        self.assertLessEqual(
                            abs(rendered.height - expected_size[1]),
                            1,
                        )
                self.assertGreater(
                    len({round(height, 1) for height in page_heights}),
                    1,
                )
                page_texts = [
                    page.extract_text() or ""
                    for page in reader.pages
                ]
                text = "\n".join(page_texts)
                self.assertIn("进攻方", text)
                self.assertIn("防守方", text)
                self.assertIn("补丁说明", text)
                self.assertIn("副喷 ✓", text)
                self.assertIn("主狙 ✓", text)
                self.assertNotIn("副喷 是", text)
                self.assertNotIn("主狙 是", text)
                self.assertNotIn("×2", text)
                self.assertIn("进攻方", page_texts[0])
                self.assertNotIn("防守方", page_texts[0])
                self.assertNotIn("补丁说明", page_texts[0])
                self.assertIn("防守方", page_texts[1])
                self.assertNotIn("进攻方", page_texts[1])
                self.assertNotIn("补丁说明", page_texts[1])
                self.assertIn("补丁说明", page_texts[2])
                self.assertNotIn("进攻方", page_texts[2])
                self.assertNotIn("防守方", page_texts[2])
                band_widths = []

                def collect_band_rectangles(operator, operands, _cm, _tm):
                    if operator != b"re" or len(operands) < 4:
                        return
                    width = float(operands[2])
                    height = float(operands[3])
                    if abs(abs(height) - 9 * mm) < 0.1:
                        band_widths.append(abs(width))

                for page in reader.pages:
                    page.extract_text(visitor_operand_before=collect_band_rectangles)
                self.assertTrue(band_widths)
                expected_width = A2[0] - 20 * mm
                for width in band_widths:
                    self.assertAlmostEqual(width, expected_width, delta=0.1)

    def test_main_reports_missing_input(self):
        with tempfile.TemporaryDirectory() as directory:
            stderr = io.StringIO()
            with contextlib.redirect_stderr(stderr):
                result = lb.main(
                    [
                        "--input",
                        str(Path(directory) / "missing.xlsx"),
                    ]
                )

        self.assertEqual(result, 1)
        self.assertIn("错误：找不到输入文件", stderr.getvalue())


class LeaderboardBatchLauncherTests(unittest.TestCase):
    def test_batch_launcher_generates_output_directory_and_preserves_exit_code(self):
        launcher = (
            Path(__file__).resolve().parents[1]
            / "run_r6_report.bat"
        )
        self.assertTrue(launcher.exists())
        content = launcher.read_text(encoding="utf-8")
        self.assertIn('cd /d "%~dp0"', content)
        self.assertIn("where python", content)
        self.assertIn("where py", content)
        self.assertIn(
            "-m r6_report.leaderboards",
            content,
        )
        self.assertIn('--inputs-dir "%~dp0inputs"', content)
        self.assertIn(
            '--input "%~dp0~temp\\r6_operator_stats.xlsx"',
            content,
        )
        self.assertIn('--output-dir "%~dp0~outputs"', content)
        self.assertIn('set "EXIT_CODE=%ERRORLEVEL%"', content)
        self.assertIn("exit /b %EXIT_CODE%", content)
        self.assertNotIn(" del ", content.lower())
        self.assertNotIn("remove-item", content.lower())


if __name__ == "__main__":
    unittest.main()
