import contextlib
import io
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from PIL import Image

import _path_setup
from r6_report import tier_chart as chart
from source_fixtures import make_report_sources


HEADERS = [
    "图标",
    "干员",
    "速度",
    "主手自动枪械（射速，发/分钟）",
    "副手自动枪械（射速，发/分钟）",
    "主狙",
    "副手霰弹",
    "次要装备",
    "Athieno评分",
]


def make_source(path, attackers, defenders, headers=HEADERS):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in (("进攻方", attackers), ("防守方", defenders)):
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
    workbook.save(path)


def source_row(
    name,
    speed=2,
    primary="R4-C（900）\nL85A2（700）",
    secondary="SMG-11（1270）",
    semiautomatic="417",
    shotgun="ITA12S",
    gadgets="手雷×2\n烟雾弹×2",
    score=100,
):
    return [None, name, speed, primary, secondary, semiautomatic, shotgun, gadgets, score]


class ParserTests(unittest.TestCase):
    def test_parses_all_named_weapons_and_matching_rates(self):
        automatic = chart.parse_automatic_weapons(
            "G8A1（850）\nAUG A2（720）\n552 Commando（690）"
        )
        named = chart.parse_named_weapons("Mk 14 EBR\nBOSG.12.2")

        self.assertEqual(
            [(item.name, item.firerate) for item in automatic],
            [
                ("G8A1", 850),
                ("AUG A2", 720),
                ("552 Commando", 690),
            ],
        )
        self.assertEqual(
            [(item.name, item.firerate) for item in named],
            [("Mk 14 EBR", None), ("BOSG.12.2", None)],
        )
        self.assertEqual(chart.parse_automatic_weapons("无自动枪械"), ())
        self.assertEqual(chart.parse_named_weapons("无"), ())

    def test_reads_only_structured_table_rows_before_trailing_status(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "source-with-status.xlsx"
            make_source(path, [source_row("Alice")], [source_row("Bob")])
            workbook = load_workbook(path)
            for side, table_name in (
                ("进攻方", "AttackOperators"),
                ("防守方", "DefenseOperators"),
            ):
                sheet = workbook[side]
                sheet.add_table(Table(displayName=table_name, ref="A1:I2"))
                sheet["A4"] = "评分来源：测试尾注"
            workbook.save(path)

            cards = chart.load_operator_cards(path)

        self.assertEqual([card.name for card in cards["进攻方"]], ["Alice"])
        self.assertEqual([card.name for card in cards["防守方"]], ["Bob"])

    def test_extracts_compact_facts_and_normalizes_gadget_aliases(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "source.xlsx"
            make_source(
                path,
                [source_row("Alice")],
                [
                    source_row(
                        "Bob",
                        speed=1,
                        primary="无自动枪械",
                        secondary="无自动枪械",
                        semiautomatic="无",
                        shotgun="无",
                        gadgets="眩晕手榴弹×3\n烟雾手榴弹×2",
                        score=0,
                    )
                ],
            )
            cards = chart.load_operator_cards(path)

        alice = cards["进攻方"][0]
        self.assertEqual(alice.primary_rpms, (900, 700))
        self.assertEqual(alice.secondary_rpms, (1270,))
        self.assertTrue(alice.has_semiautomatic)
        self.assertTrue(alice.has_secondary_shotgun)
        self.assertEqual(
            [(item.name, item.firerate) for item in alice.primary_weapons],
            [("R4-C", 900), ("L85A2", 700)],
        )
        self.assertEqual(
            [(item.name, item.firerate) for item in alice.secondary_weapons],
            [("SMG-11", 1270)],
        )
        self.assertEqual(
            [item.name for item in alice.semiautomatic_weapons],
            ["417"],
        )
        self.assertEqual(
            [item.name for item in alice.secondary_shotguns],
            ["ITA12S"],
        )
        self.assertEqual(
            alice.gadgets,
            (
                chart.GadgetItem("破片手榴弹", 2),
                chart.GadgetItem("烟雾弹", 2),
            ),
        )
        self.assertEqual((alice.tier, alice.source_order), ("S", 0))

        bob = cards["防守方"][0]
        self.assertEqual(bob.primary_rpms, ())
        self.assertEqual(bob.secondary_rpms, ())
        self.assertFalse(bob.has_semiautomatic)
        self.assertFalse(bob.has_secondary_shotgun)
        self.assertEqual(
            bob.gadgets,
            (
                chart.GadgetItem("闪光弹", 3),
                chart.GadgetItem("烟雾弹", 2),
            ),
        )
        self.assertEqual(bob.tier, "F")

    def test_extract_rpms_preserves_display_order_and_rounds_whole_values(self):
        self.assertEqual(chart.extract_rpms("A（900）\nB（700.0）"), (900, 700))
        self.assertEqual(chart.extract_rpms("无自动枪械"), ())
        with self.assertRaisesRegex(chart.TierChartError, "无法提取射速"):
            chart.extract_rpms("R4-C")

    def test_rejects_missing_headers_duplicate_names_unknown_scores_and_bad_speed(self):
        cases = [
            (HEADERS[:-1], [source_row("Alice")[:-1]], "缺少必需表头"),
            (HEADERS, [source_row("Alice"), source_row("Alice")], "重复干员.*Alice"),
            (HEADERS, [source_row("Alice", score=99)], "未知评分.*Alice"),
            (HEADERS, [source_row("Alice", speed="快")], "速度不是数字.*Alice"),
        ]
        for headers, rows, message in cases:
            with self.subTest(message=message), tempfile.TemporaryDirectory() as directory:
                path = Path(directory) / "source.xlsx"
                make_source(path, rows, [source_row("Bob", score=85)], headers=headers)
                with self.assertRaisesRegex(chart.TierChartError, message):
                    chart.load_operator_cards(path)


class GadgetIconTests(unittest.TestCase):
    def test_prepares_all_weapon_icons_as_cropped_dark_silhouettes(self):
        source = io.BytesIO()
        image = Image.new("RGBA", (12, 8), (255, 255, 255, 0))
        for x in range(3, 9):
            for y in range(2, 6):
                image.putpixel((x, y), (240, 120, 60, 255))
        image.save(source, format="PNG")
        calls = []

        def query_json(parameters):
            title = parameters["titles"]
            return {
                "query": {
                    "pages": [{
                        "imageinfo": [{
                            "url": (
                                "https://huiji-public.huijistatic.com/r6s/"
                                + title.rsplit(" ", 1)[-1].replace(".png", "")
                                + ".png"
                            )
                        }]
                    }]
                }
            }

        def runner(command, **kwargs):
            calls.append(command)
            destination = Path(command[command.index("--output") + 1])
            destination.write_bytes(source.getvalue())
            return SimpleNamespace()

        items = (
            chart.WeaponItem("R4-C", "r4-c", 860),
            chart.WeaponItem("G8A1", "g8a1", 850),
            chart.WeaponItem("R4-C", "r4-c", 860),
        )
        with tempfile.TemporaryDirectory() as directory:
            paths = chart.prepare_weapon_icons(
                items,
                Path(directory),
                query_json=query_json,
                run_command=runner,
                which=lambda _: "C:/curl.exe",
                sleep=lambda _: None,
            )
            self.assertEqual(tuple(paths), ("r4-c", "g8a1"))
            self.assertEqual(len(calls), 2)
            with Image.open(paths["r4-c"]) as icon:
                self.assertEqual(icon.size, (6, 4))
                visible = [
                    pixel
                    for pixel in icon.convert("RGBA").get_flattened_data()
                    if pixel[3] > 0
                ]

        self.assertTrue(visible)
        self.assertTrue(
            all(pixel[:3] == (32, 35, 39) for pixel in visible)
        )

    def test_official_gadget_file_map_is_complete_and_exact(self):
        self.assertEqual(
            chart.GADGET_FILES,
            {
                "倒刺铁丝网": "文件:R6S gp Barbed wire.png",
                "冲击手榴弹": "文件:R6S gp Impact Grenade.png",
                "感应警报器": "文件:R6S gp Proximity Alarm.png",
                "破片手榴弹": "文件:R6S gp Frag Grenade.png",
                "机动护盾": "文件:R6S gp Deployable Shield.png",
                "烟雾弹": "文件:R6S gp Smoke Grenade.png",
                "爆破炸药": "文件:R6S gp Breach Charge.png",
                "电磁脉冲式冲击弹": "文件:R6S gp Impact emp Grenade.png",
                "闪光弹": "文件:R6S gp Stun Grenade.png",
                "硬突破炸药": "文件:R6S gp SecondaryBreacher.png",
                "观测工具阻拦器": "文件:R6S gp Observation Blocker.png",
                "遥控炸药": "文件:R6S gp Nitro Cell.png",
                "阔剑地雷": "文件:R6S gp Claymore.png",
                "防弹摄像头": "文件:R6S gp Bulletproof camera.png",
            },
        )

    def test_resolves_only_official_huiji_imageinfo_urls(self):
        expected_url = "https://huiji-public.huijistatic.com/r6s/example.png"
        calls = []

        def query_json(parameters):
            calls.append(parameters)
            return {"query": {"pages": [{"imageinfo": [{"url": expected_url}]}]}}

        self.assertEqual(
            chart.resolve_wiki_file_url("文件:R6S gp Frag Grenade.png", query_json),
            expected_url,
        )
        self.assertEqual(calls[0]["prop"], "imageinfo")
        self.assertEqual(calls[0]["titles"], "文件:R6S gp Frag Grenade.png")

        bad_responses = [
            {"query": {"pages": [{}]}},
            {"query": {"pages": [{"imageinfo": [{"url": "https://example.com/a.png"}]}]}},
        ]
        for response in bad_responses:
            with self.subTest(response=response):
                with self.assertRaisesRegex(chart.TierChartError, "图标地址"):
                    chart.resolve_wiki_file_url("文件:R6S gp Frag Grenade.png", lambda _: response)

    def test_prepares_valid_png_and_reuses_cache(self):
        image_bytes = io.BytesIO()
        Image.new("RGBA", (24, 24), "red").save(image_bytes, format="PNG")
        calls = []

        def runner(command, **kwargs):
            calls.append(command)
            destination = Path(command[command.index("--output") + 1])
            destination.write_bytes(image_bytes.getvalue())
            return SimpleNamespace()

        with tempfile.TemporaryDirectory() as directory:
            icon_dir = Path(directory)
            icons = chart.prepare_gadget_icons(
                (chart.GadgetItem("破片手榴弹", 2),),
                icon_dir,
                query_json=lambda _: {
                    "query": {
                        "pages": [
                            {
                                "imageinfo": [
                                    {
                                        "url": "https://huiji-public.huijistatic.com/r6s/frag.png"
                                    }
                                ]
                            }
                        ]
                    }
                },
                run_command=runner,
                which=lambda _: "C:/curl.exe",
                sleep=lambda _: None,
            )
            again = chart.prepare_gadget_icons(
                (chart.GadgetItem("破片手榴弹", 2),),
                icon_dir,
                query_json=lambda _: self.fail("valid cache should skip API"),
                run_command=lambda *args, **kwargs: self.fail("valid cache should skip curl"),
                which=lambda _: "C:/curl.exe",
                sleep=lambda _: None,
            )

            self.assertEqual(icons, again)
            self.assertTrue(icons["破片手榴弹"].is_file())
            self.assertEqual(len(calls), 1)

    def test_impact_emp_uses_official_line_art_source(self):
        image_bytes = io.BytesIO()
        source = Image.new("RGBA", (16, 16), (0, 0, 0, 0))
        source.putpixel((7, 7), (255, 255, 255, 255))
        source.putpixel((8, 8), (255, 255, 255, 128))
        source.save(image_bytes, format="PNG")
        calls = []

        def runner(command, **kwargs):
            calls.append(command)
            destination = Path(command[command.index("--output") + 1])
            destination.write_bytes(image_bytes.getvalue())
            return SimpleNamespace()

        with tempfile.TemporaryDirectory() as directory:
            icons = chart.prepare_gadget_icons(
                (chart.GadgetItem("电磁脉冲式冲击弹", 2),),
                Path(directory),
                query_json=lambda _: self.fail(
                    "Impact EMP should use the Ubisoft source"
                ),
                run_command=runner,
                which=lambda _: "C:/curl.exe",
                sleep=lambda _: None,
            )

            self.assertEqual(
                calls[0][-1],
                (
                    "https://staticctf.ubisoft.com/"
                    "J3yJr34U2pZ2Ieem48Dwy9uqj5PNUQTn/"
                    "7izurbA5jDmnsmdeBdgKZO/"
                    "29bca81243dda4084a92521ac0c03592/"
                    "R6S-EMP-Impact-grenade.png"
                ),
            )
            with Image.open(icons["电磁脉冲式冲击弹"]) as icon:
                visible = [
                    pixel
                    for pixel in icon.convert("RGBA").get_flattened_data()
                    if pixel[3] > 0
                ]

        self.assertTrue(visible)
        self.assertTrue(
            all(pixel[:3] == (0, 0, 0) for pixel in visible)
        )

    def test_rejects_unknown_gadget_and_retries_invalid_download_four_times(self):
        with tempfile.TemporaryDirectory() as directory:
            with self.assertRaisesRegex(chart.TierChartError, "未知次要装备.*未知装备"):
                chart.prepare_gadget_icons(
                    (chart.GadgetItem("未知装备", 1),), Path(directory)
                )

        calls = []

        def runner(command, **kwargs):
            calls.append(command)
            Path(command[command.index("--output") + 1]).write_bytes(b"not a png")
            return SimpleNamespace()

        with tempfile.TemporaryDirectory() as directory:
            with self.assertRaisesRegex(chart.TierChartError, "4 次尝试"):
                chart.prepare_gadget_icons(
                    (chart.GadgetItem("烟雾弹", 2),),
                    Path(directory),
                    query_json=lambda _: {
                        "query": {
                            "pages": [
                                {
                                    "imageinfo": [
                                        {
                                            "url": "https://huiji-public.huijistatic.com/r6s/smoke.png"
                                        }
                                    ]
                                }
                            ]
                        }
                    },
                    run_command=runner,
                    which=lambda _: "C:/curl.exe",
                    sleep=lambda _: None,
                )
        self.assertEqual(len(calls), 4)

    def test_gadget_token_crops_transparent_padding_before_scaling(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "padded.png"
            output = root / "token.png"
            image = Image.new("RGBA", (100, 100), (0, 0, 0, 0))
            for x in range(45, 55):
                for y in range(45, 55):
                    image.putpixel((x, y), (255, 0, 0, 255))
            image.save(source)

            chart._draw_gadget_token(source, None, output)
            token = Image.open(output).convert("RGBA")
            red_x = [
                x
                for x in range(token.width)
                for y in range(token.height)
                if token.getpixel((x, y))[0] > 200 and token.getpixel((x, y))[3] > 0
            ]

        self.assertGreaterEqual(max(red_x) - min(red_x) + 1, 16)


def make_card(name, side, order, tier="S", score=100):
    return chart.OperatorCard(
        side=side,
        name=name,
        speed=2,
        score=score,
        tier=tier,
        primary_rpms=(900, 700),
        secondary_rpms=(1270,),
        has_semiautomatic=True,
        has_secondary_shotgun=True,
        gadgets=(
            chart.GadgetItem("破片手榴弹", 2),
            chart.GadgetItem("烟雾弹", 2),
        ),
        source_order=order,
    )


class WorkbookLayoutTests(unittest.TestCase):
    def test_renders_two_five_card_tier_sheets_with_images_and_print_settings(self):
        cards = {
            "进攻方": [make_card(f"Attacker {number}", "进攻方", number) for number in range(1, 7)],
            "防守方": [make_card("Defender", "防守方", 1, tier="A", score=85)],
        }
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            badge_dir = root / "badges"
            gadget_dir = root / "gadgets"
            badge_dir.mkdir()
            gadget_dir.mkdir()
            for card in cards["进攻方"] + cards["防守方"]:
                Image.new("RGBA", (64, 64), "red").save(
                    badge_dir / f"{chart.operator_key(card.name)}.png"
                )
            gadget_icons = {}
            for name in ("破片手榴弹", "烟雾弹"):
                path = gadget_dir / f"{name}.png"
                Image.new("RGBA", (32, 32), "blue").save(path)
                gadget_icons[name] = path

            output = root / "tier.xlsx"
            chart.write_tier_workbook(
                output,
                cards,
                badge_dir,
                gadget_icons,
                make_report_sources(),
            )
            workbook = load_workbook(output)

        self.assertEqual(
            workbook.sheetnames,
            ["进攻方简图", "防守方简图", "补丁说明"],
        )
        attackers = workbook["进攻方简图"]
        defenders = workbook["防守方简图"]
        self.assertIn("A3:A6", {str(item) for item in attackers.merged_cells.ranges})
        self.assertEqual(
            [
                value
                for value in (
                    attackers.cell(row, 1).value
                    for row in range(3, attackers.max_row + 1)
                )
                if value in chart.TIER_ORDER
            ],
            ["S", "A", "B", "C", "D", "F"],
        )
        self.assertEqual(attackers["C3"].value, "Attacker 1\n2速 · 主 900/700 · 副 1270\n主狙 ✓ · 副喷 ✓")
        self.assertEqual(attackers["C5"].value, "Attacker 6\n2速 · 主 900/700 · 副 1270\n主狙 ✓ · 副喷 ✓")
        self.assertEqual(len(attackers._images), 6 + (6 * 2))
        self.assertEqual(len(defenders._images), 1 + 2)
        for sheet in (attackers, defenders):
            self.assertEqual(sheet.sheet_properties.pageSetUpPr.fitToPage, True)
            self.assertEqual(sheet.page_setup.orientation, "landscape")
            self.assertEqual(sheet.page_setup.fitToWidth, 1)
            self.assertEqual(sheet.page_setup.fitToHeight, 1)
            self.assertFalse(sheet.sheet_view.showGridLines)
            self.assertEqual(sheet.freeze_panes, "B3")
            self.assertTrue(sheet.print_area)
            self.assertIn(
                "补丁区间：", sheet.cell(sheet.max_row, 1).value
            )
            self.assertIn(
                f"A{sheet.max_row}:K{sheet.max_row}",
                {str(item) for item in sheet.merged_cells.ranges},
            )
            self.assertIn(str(sheet.max_row), str(sheet.print_area))
        self.assertEqual(
            workbook["补丁说明"]["A1"].value,
            "Y11S2 视频评分后续补丁说明",
        )

    def test_rejects_missing_operator_badge_and_unknown_card_tier(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            root.joinpath("badges").mkdir()
            gadget = root / "frag.png"
            Image.new("RGBA", (20, 20), "blue").save(gadget)
            icons = {"破片手榴弹": gadget, "烟雾弹": gadget}
            with self.assertRaisesRegex(chart.TierChartError, "找不到干员 Badge.*Alice"):
                chart.write_tier_workbook(
                    root / "missing.xlsx",
                    {"进攻方": [make_card("Alice", "进攻方", 1)], "防守方": []},
                    root / "badges",
                    icons,
                    make_report_sources(),
                )

            bad = make_card("Alice", "进攻方", 1, tier="X")
            Image.new("RGBA", (64, 64), "red").save(
                root / "badges" / f"{chart.operator_key('Alice')}.png"
            )
            with self.assertRaisesRegex(chart.TierChartError, "未知阶层.*Alice"):
                chart.write_tier_workbook(
                    root / "bad-tier.xlsx",
                    {"进攻方": [bad], "防守方": []},
                    root / "badges",
                    icons,
                    make_report_sources(),
                )


class CliTests(unittest.TestCase):
    def test_main_generates_workbook_and_prints_counts_and_absolute_path(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "source.xlsx"
            output = root / "tier.xlsx"
            badge_dir = root / "badges"
            gadget_dir = root / "gadgets"
            badge_dir.mkdir()
            gadget_dir.mkdir()
            make_source(
                source,
                [source_row("Alice")],
                [source_row("Bob", score=85)],
            )
            for name in ("Alice", "Bob"):
                Image.new("RGBA", (64, 64), "red").save(
                    badge_dir / f"{chart.operator_key(name)}.png"
                )
            gadget_icons = {}
            for name in ("破片手榴弹", "烟雾弹"):
                path = gadget_dir / f"{name}.png"
                Image.new("RGBA", (32, 32), "blue").save(path)
                gadget_icons[name] = path

            stdout = io.StringIO()
            stderr = io.StringIO()
            with contextlib.redirect_stdout(stdout), contextlib.redirect_stderr(stderr):
                result = chart.main(
                    [
                        "--input",
                        str(source),
                        "--output",
                        str(output),
                        "--icons-dir",
                        str(badge_dir),
                        "--gadget-icons-dir",
                        str(gadget_dir),
                    ],
                    gadget_icon_preparer=lambda items, directory: gadget_icons,
                    source_loader=lambda path: make_report_sources(),
                )

            self.assertEqual(result, 0)
            self.assertTrue(output.is_file())
            self.assertEqual(stderr.getvalue(), "")
            self.assertIn("进攻方干员：1", stdout.getvalue())
            self.assertIn("防守方干员：1", stdout.getvalue())
            self.assertIn(str(output.resolve()), stdout.getvalue())

    def test_main_reports_missing_input_on_stderr(self):
        stderr = io.StringIO()
        with tempfile.TemporaryDirectory() as directory, contextlib.redirect_stderr(stderr):
            result = chart.main(["--input", str(Path(directory) / "missing.xlsx")])
        self.assertEqual(result, 1)
        self.assertIn("错误：找不到输入文件", stderr.getvalue())

    def test_script_entrypoint_defines_all_helpers_before_calling_main(self):
        project_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "source.xlsx"
            make_source(source, [source_row("Alice")], [source_row("Bob", score=85)])
            result = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "r6_report.tier_chart",
                    "--input",
                    str(source),
                    "--output",
                    str(root / "output.xlsx"),
                    "--icons-dir",
                    str(root / "missing-badges"),
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                check=False,
                cwd=project_root,
                env={
                    **os.environ,
                    "PYTHONIOENCODING": "utf-8",
                    "PYTHONPATH": str(project_root / "src"),
                },
            )
        self.assertEqual(result.returncode, 1)
        self.assertNotIn("NameError", result.stderr)
        self.assertIn("错误：找不到干员 Badge", result.stderr)


if __name__ == "__main__":
    unittest.main()
